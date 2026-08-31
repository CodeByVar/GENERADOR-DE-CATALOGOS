# -*- coding: utf-8 -*-
"""
DongCheng Herramientas - Generador de Catálogo con HTML + CSS
===========================================================
REQUISITO: pip install openpyxl pillow
USO:
  1. Pon este script en la misma carpeta que catalogos.xlsx
  2. En Vista_Catalogo columna J desde fila 4 escribe los códigos
  3. Guarda y cierra el Excel
  4. Ejecuta el servidor web mediante Generar_Catalogo.bat
"""

from openpyxl import load_workbook
from datetime import date
from io import BytesIO
from PIL import Image as PILImage
import os, sys, math, re
import urllib.request
import subprocess
import hashlib
import json
import time

# ─── CONFIGURACIÓN ────────────────────────────────────────────
# La base de datos Excel se guarda como caché interna en temp_imgs
# para evitar tener el archivo catalogos.xlsx visible en la carpeta principal.
ARCHIVO_EXCEL       = os.path.join("temp_imgs", "catalogos_db_cache.xlsx")
URL_GOOGLE_SHEETS   = "https://docs.google.com/spreadsheets/d/181FkDYPFME5Fx75og4tNO3mvBMSGDY-M9IxGFcR28SI/edit?usp=sharing"
URL_STOCK_API       = "https://script.google.com/macros/s/AKfycbxrXCYxH9JX-uO2rw5Wg7XY5PnbKso50ugmpkTnrPacwy12GoMpxn-AvlbRZ_m0a9k45w/exec"
HOJA_DB             = "FORMATO INVENTARIO"
HOJA_VISTA          = "Vista_Catalogo"
HOJA_CATALOGO       = "CATALOGO"
COLUMNA_CODIGOS     = 10   # columna J en Vista_Catalogo
FILA_INICIO_CODIGOS = 4

# Columnas en DC MANU
COL_CATEGORIA = 2   # B
COL_CODIGO    = 3   # C
COL_IMAGEN    = 4   # D  ← imágenes están en esta columna
COL_TIPO      = 5   # E
COL_NOMBRE    = 6   # F
COL_SIZE      = 7   # G
COL_DETALLE   = 8   # H
COL_UNI       = 9   # I

FILA_INICIO_DB = 4
# ──────────────────────────────────────────────────────────────

def extraer_id_google_sheets(url_or_id):
    if not url_or_id:
        return None
    if "/" not in url_or_id and len(url_or_id) > 20:
        return url_or_id
    match = re.search(r"/d/([a-zA-Z0-9-_]+)", url_or_id)
    if match:
        return match.group(1)
    return None

def descargar_base_de_datos_nube(url=None, destino=ARCHIVO_EXCEL):
    """
    Descarga la versión más reciente del Excel desde Google Sheets / Google Drive.
    """
    input_url = url or URL_GOOGLE_SHEETS
    if not input_url:
        return False
        
    doc_id = extraer_id_google_sheets(input_url)
    if doc_id:
        download_url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=xlsx&t={int(time.time())}"
    else:
        download_url = input_url
 
    print(f"\n[NUBE] Sincronizando base de datos desde Google Drive...")
    try:
        req = urllib.request.Request(
            download_url,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        )
        with urllib.request.urlopen(req, timeout=30) as response:
            data = response.read()
            if len(data) > 1000:  # Verificación mínima de tamaño
                dir_name = os.path.dirname(destino)
                if dir_name and not os.path.exists(dir_name):
                    os.makedirs(dir_name)
                with open(destino, "wb") as out_file:
                    out_file.write(data)
                print(f"[NUBE] [OK] Base de datos actualizada con éxito desde Google Sheets ({len(data)} bytes)")
                return True
            else:
                print(f"[NUBE] [AVISO] La respuesta descargada fue muy pequeña ({len(data)} bytes).")
                return False
    except urllib.error.HTTPError as he:
        if he.code == 404 or he.code == 403:
            print(f"[NUBE] [AVISO] El documento en Google Drive está privado o restringido (HTTP {he.code}).")
            print(f"[NUBE] Para activar la sincronización automática, haz clic en 'Compartir' en Google Sheets y selecciona 'Cualquier persona con el enlace'.")
        else:
            print(f"[NUBE] [AVISO] Error HTTP {he.code} al descargar de la nube: {he}")
        print(f"[NUBE] Se utilizará la copia local existente de '{destino}'.")
        return False
    except Exception as e:
        print(f"[NUBE] [AVISO] No se pudo descargar desde la nube: {e}")
        print(f"[NUBE] Se utilizará la copia local existente de '{destino}'.")
        return False

# Temas y logotipos para las marcas
BRAND_THEMES = {
    "DONGCHENG ELECT.": {
        "logo": "Logo DonElec.webp",
        "display_name": "DONGCHENG ELECTRICO",
        "header_bg": "FFFFFF",
        "subtitle_color": "005BAC",
        "category_bg": "DBEAFE",
        "category_fg": "1E40AF",
        "card_header_bg": "005BAC",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "E2E8F0",
        "card_measure_fg": "000000",
    },
    "DONGCHENG MANUAL": {
        "logo": "Logo DonManu.jpg",
        "display_name": "DONGCHENG MANUAL",
        "header_bg": "FFFFFF",
        "subtitle_color": "334155",
        "category_bg": "E2E8F0",
        "category_fg": "334155",
        "card_header_bg": "005BAC",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "E2E8F0",
        "card_measure_fg": "000000",
    },
    "FERRAWYY": {
        "logo": "Logo Ferr.png",
        "display_name": "FERRAWYY",
        "header_bg": "F97316",
        "subtitle_color": "FFFFFF",
        "category_bg": "FFEDD5",
        "category_fg": "C2410C",
        "card_header_bg": "EA580C",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FFEDD5",
        "card_measure_fg": "C2410C",
    },
    "UYUSTOOLS ELECT.": {
        "logo": "Logo UyuElec.jpg",
        "display_name": "UYUSTOOLS ELECTRICO",
        "header_bg": "FDC800",
        "subtitle_color": "000000",
        "category_bg": "1E293B",
        "category_fg": "FFFFFF",
        "card_header_bg": "FDC800",
        "card_header_fg": "000000",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEF3C7",
        "card_measure_fg": "92400E",
    },
    "UYUSTOOLS MANUAL": {
        "logo": "Logo UyuManu.png",
        "display_name": "UYUSTOOLS MANUAL",
        "header_bg": "FDC800",
        "subtitle_color": "000000",
        "category_bg": "334155",
        "category_fg": "FFFFFF",
        "card_header_bg": "FDC800",
        "card_header_fg": "000000",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEF3C7",
        "card_measure_fg": "92400E",
    },
    "PEGASUS": {
        "logo": "Logo Pega.png",
        "display_name": "PEGASUS",
        "header_bg": "FFFFFF",
        "subtitle_color": "1E293B",
        "category_bg": "F1F5F9",
        "category_fg": "0F172A",
        "card_header_bg": "334155",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "E2E8F0",
        "card_measure_fg": "0F172A",
    },
    "FERTON": {
        "logo": "Logo Fer.jpg",
        "display_name": "FERTON",
        "header_bg": "000000",
        "subtitle_color": "FBBF24",
        "category_bg": "FEE2E2",
        "category_fg": "991B1B",
        "card_header_bg": "1E293B",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEF3C7",
        "card_measure_fg": "78350F",
    },
    "AQUASTRONG": {
        "logo": "Logo Aquas.webp",
        "display_name": "AQUASTRONG",
        "header_bg": "FFFFFF",
        "subtitle_color": "0284C7",
        "category_bg": "E0F2FE",
        "category_fg": "0369A1",
        "card_header_bg": "0284C7",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "E0F2FE",
        "card_measure_fg": "0369A1",
    },
    "GALAXIA": {
        "logo": "Logo Galaxia.png",
        "display_name": "GALAXIA",
        "header_bg": "FFFFFF",
        "subtitle_color": "16355F",
        "category_bg": "E0E7FF",
        "category_fg": "16355F",
        "card_header_bg": "16355F",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "E0E7FF",
        "card_measure_fg": "16355F",
    },
    "GATE": {
        "logo": "Logo Gate.png",
        "display_name": "GATE",
        "header_bg": "FFFFFF",
        "subtitle_color": "EC403E",
        "category_bg": "FFECEC",
        "category_fg": "991B1B",
        "card_header_bg": "EC403E",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEE2E2",
        "card_measure_fg": "991B1B",
    },
    "CROWN": {
        "logo": "Logo Crown.png",
        "display_name": "CROWN",
        "header_bg": "FFFFFF",
        "subtitle_color": "CA0A10", # Official Crown red
        "category_bg": "FFECEC", # Light red/rose matching brand
        "category_fg": "CA0A10", # Official Crown red
        "card_header_bg": "CA0A10", # Official Crown red
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "E5E7EB", # Light gray matching screenshot
        "card_measure_fg": "CA0A10", # Official Crown red
    },
    "RIO": {
        "logo": "Logo Rio.png",
        "display_name": "RIO",
        "header_bg": "FFFFFF",
        "subtitle_color": "9B2226", # True tinto red
        "category_bg": "FEE2E2",
        "category_fg": "9B2226", # True tinto red
        "card_header_bg": "9B2226", # True tinto red
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FFE4E6",
        "card_measure_fg": "9B2226", # True tinto red
    },
    "OMEGA": {
        "logo": "Logo Omega.png",
        "display_name": "OMEGA",
        "header_bg": "FFFFFF",
        "subtitle_color": "7A1C1C", # Burgundy red (rojo tinto)
        "category_bg": "FEE2E2", # Light burgundy pink
        "category_fg": "7A1C1C", # Burgundy red (rojo tinto)
        "card_header_bg": "7A1C1C", # Burgundy red (rojo tinto)
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "E0F2FE", # Light blue
        "card_measure_fg": "0284C7", # Blue
    },
    "TOTAL": {
        "logo": "Logo total.png",
        "display_name": "TOTAL",
        "header_bg": "FFFFFF",
        "subtitle_color": "186E6F", # Characteristic green/teal
        "category_bg": "E6F3F3", # Light teal/green background matching brand
        "category_fg": "186E6F", # Characteristic green/teal
        "card_header_bg": "186E6F", # Characteristic green/teal
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FFECEC", # Light red/rose background
        "card_measure_fg": "E2231A", # Brand red text
    },
    "WADFOW": {
        "logo": "Logo Wadfow.png",
        "display_name": "WADFOW",
        "header_bg": "FFFFFF",
        "subtitle_color": "0A4E9D", # Wadfow blue
        "category_bg": "EFF6FF", # Very light blue background
        "category_fg": "0A4E9D", # Wadfow blue
        "card_header_bg": "0A4E9D", # Wadfow blue
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FFEDD5", # Light orange background
        "card_measure_fg": "FD5F00", # Wadfow orange text
    },
    "NEVA": {
        "logo": "LogoNeva.png",
        "display_name": "NEVA",
        "header_bg": "FFFFFF",
        "subtitle_color": "000000",
        "category_bg": "FEF08A", # Light yellow (Yellow 200)
        "category_fg": "000000",
        "card_header_bg": "000000", # Black
        "card_header_fg": "FACC15", # Yellow (Yellow 400)
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEF9C3", # Very light yellow (Yellow 100)
        "card_measure_fg": "000000",
    },
    "LUTIAN": {
        "logo": "LogoLutian.png",
        "display_name": "LUTIAN",
        "header_bg": "FFFFFF",
        "subtitle_color": "4F8A10", # Dark green text
        "category_bg": "F7FEE7", # Light lime green background (Lime 50)
        "category_fg": "3F6212", # Dark lime green text (Lime 800)
        "card_header_bg": "84CC16", # Lime green (Lime 500)
        "card_header_fg": "FFFFFF",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "F1FEE7", # Very light lime green
        "card_measure_fg": "4D7C0F", # Lime green text (Lime 700)
    },
    "DWT": {
        "logo": "Logo Dwt.png",
        "display_name": "DWT",
        "header_bg": "FFFFFF",
        "subtitle_color": "E11D24", # Red branding
        "category_bg": "FEE2E2",
        "category_fg": "991B1B",
        "card_header_bg": "E11D24",
        "card_header_fg": "FFFFFF",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEE2E2",
        "card_measure_fg": "991B1B",
    },
    "LION": {
        "logo": "Logo Lion.png",
        "display_name": "LION",
        "header_bg": "FFFFFF",
        "subtitle_color": "B45309", # Warm amber/gold
        "category_bg": "FEF3C7", # Light amber/yellow
        "category_fg": "92400E", # Dark amber
        "card_header_bg": "FACC15", # Brand Yellow
        "card_header_fg": "000000", # High contrast black text
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEF3C7",
        "card_measure_fg": "92400E",
    },
    "NORSTAR": {
        "logo": "Logo Norstar.png",
        "display_name": "NORSTAR",
        "header_bg": "FFFFFF",
        "subtitle_color": "D97706", # Norstar Ochre / Gold
        "category_bg": "FEF3C7",
        "category_fg": "92400E",
        "card_header_bg": "D97706",
        "card_header_fg": "FFFFFF",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEF9C3",
        "card_measure_fg": "713F12",
    },
    "POWERMAQ": {
        "logo": "Logo Powermaq.png",
        "display_name": "POWERMAQ",
        "header_bg": "FFFFFF",
        "subtitle_color": "1E293B", # Slate / Graphite
        "category_bg": "F1F5F9",
        "category_fg": "0F172A",
        "card_header_bg": "1E293B",
        "card_header_fg": "FFFFFF",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "E2E8F0",
        "card_measure_fg": "0F172A",
    },
    "STANFORD": {
        "logo": "Logo Stanford.png",
        "display_name": "STANFORD",
        "header_bg": "FFFFFF",
        "subtitle_color": "EA580C", # Signature Stanford Orange
        "category_bg": "FFEDD5", # Light orange
        "category_fg": "9A3412", # Dark orange
        "card_header_bg": "EA580C",
        "card_header_fg": "FFFFFF",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FFEDD5",
        "card_measure_fg": "C2410C",
    },
    "WYYMET": {
        "logo": "Logo Wyymet.png",
        "display_name": "WYYMET",
        "header_bg": "FFFFFF",
        "subtitle_color": "B91C1C", # Crimson / Dark Red
        "category_bg": "FEE2E2",
        "category_fg": "991B1B",
        "card_header_bg": "B91C1C",
        "card_header_fg": "FFFFFF",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEE2E2",
        "card_measure_fg": "991B1B",
    },
    "WYMET": {
        "logo": "Logo Wyymet.png",
        "display_name": "WYYMET",
        "header_bg": "FFFFFF",
        "subtitle_color": "B91C1C",
        "category_bg": "FEE2E2",
        "category_fg": "991B1B",
        "card_header_bg": "B91C1C",
        "card_header_fg": "FFFFFF",
        "card_detail_bg": "FFFFFF",
        "card_photo_bg": "FFFFFF",
        "card_measure_bg": "FEE2E2",
        "card_measure_fg": "991B1B",
    }
}

def convert_svg_to_png(svg_path, png_path):
    try:
        from svglib.svglib import svg2rlg
        from reportlab.graphics import renderPM
    except ImportError:
        try:
            print("  Instalando librerías necesarias para procesar archivos SVG (svglib y reportlab)...")
            subprocess.run([sys.executable, "-m", "pip", "install", "svglib", "reportlab"], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            from svglib.svglib import svg2rlg
            from reportlab.graphics import renderPM
        except Exception as e:
            print(f"  [AVISO] No se pudo instalar svglib/reportlab: {e}.")
            return False
    try:
        drawing = svg2rlg(svg_path)
        renderPM.drawToFile(drawing, png_path, fmt="PNG")
        return True
    except Exception as e:
        print(f"  [AVISO] Error al convertir SVG a PNG: {e}")
        return False

def get_brand_theme(tipo):
    tipo_norm = str(tipo).upper().strip()
    for key, theme in BRAND_THEMES.items():
        if key.upper().replace(" ", "").replace(".", "") in tipo_norm.replace(" ", "").replace(".", ""):
            logo_path = theme["logo"]
            if logo_path:
                if os.path.exists(logo_path):
                    if logo_path.lower().endswith(".svg"):
                        png_path = logo_path[:-4] + ".png"
                        if convert_svg_to_png(logo_path, png_path):
                            theme_copy = theme.copy()
                            theme_copy["logo"] = png_path
                            return theme_copy
                    return theme
                base_name = os.path.splitext(logo_path)[0]
                for f in os.listdir("."):
                    if f.lower().startswith(base_name.lower()):
                        target_logo = f
                        if f.lower().endswith(".svg"):
                            png_path = base_name + ".png"
                            if convert_svg_to_png(f, png_path):
                                target_logo = png_path
                            else:
                                continue
                        theme_copy = theme.copy()
                        theme_copy["logo"] = target_logo
                        return theme_copy
            return theme
            
    clean_tipo = tipo_norm.replace(" ", "").replace(".", "").replace("-", "")
    for f in os.listdir("."):
        if f.lower().startswith("logo"):
            name_part = os.path.splitext(f.lower()[4:])[0].strip().replace(" ", "").replace("-", "").replace(".", "")
            if name_part and (name_part in clean_tipo.lower() or clean_tipo.lower() in name_part):
                if "ferr" in name_part and "ferton" in clean_tipo.lower():
                    continue
                target_logo = f
                if f.lower().endswith(".svg"):
                    base_name = os.path.splitext(f)[0]
                    png_path = base_name + ".png"
                    if convert_svg_to_png(f, png_path):
                        target_logo = png_path
                    else:
                        continue
                return {
                    "logo": target_logo,
                    "display_name": tipo_norm,
                    "header_bg": "FFFFFF",
                    "subtitle_color": "0F172A",
                    "category_bg": "F1F5F9",
                    "category_fg": "0F172A",
                    "card_header_bg": "334155",
                }
                
    return {
        "logo": None,
        "display_name": tipo_norm,
        "header_bg": "1E293B",
        "subtitle_color": "FFFFFF",
        "category_bg": "0F172A",
        "category_fg": "FFFFFF",
        "card_header_bg": "334155",
    }

def autocrop_image(img, tolerance=20):
    from PIL import ImageChops
    if img.mode != 'RGBA':
        img = img.convert('RGBA')
    bg_color = img.getpixel((0, 0))
    bg = PILImage.new("RGBA", img.size, bg_color)
    diff = ImageChops.difference(img, bg)
    diff_gray = diff.convert("L")
    threshold_img = diff_gray.point(lambda p: 255 if p > tolerance else 0)
    bbox = threshold_img.getbbox()
    if bbox:
        left, top, right, bottom = bbox
        left = max(0, left - 6)
        top = max(0, top - 6)
        right = min(img.width, right + 6)
        bottom = min(img.height, bottom + 6)
        return img.crop((left, top, right, bottom))
    return img

def generar_logo_blanco_empresa(input_path, output_path):
    """
    Convierte el logo de la empresa a una silueta puramente blanca con fondo transparente.
    """
    if os.path.exists(output_path) and os.path.getmtime(input_path) <= os.path.getmtime(output_path):
        return output_path
    try:
        img = PILImage.open(input_path).convert("RGBA")
        
        # Recortar márgenes iniciales
        bg_color = img.getpixel((0, 0))
        from PIL import ImageChops
        bg = PILImage.new("RGBA", img.size, bg_color)
        diff = ImageChops.difference(img, bg)
        diff_gray = diff.convert("L")
        threshold_img = diff_gray.point(lambda p: 255 if p > 30 else 0)
        bbox = threshold_img.getbbox()
        if bbox:
            img = img.crop(bbox)
            
        # Convertir elementos de color a blanco y hacer el fondo negro transparente
        datas = img.getdata()
        new_data = []
        for item in datas:
            r, g, b, a = item
            brightness = (r + g + b) / 3
            if brightness < 65:  # Umbral para el fondo oscuro
                new_data.append((0, 0, 0, 0))
            else:
                new_data.append((255, 255, 255, 255))
        img.putdata(new_data)
        
        # Volver a recortar los márgenes transparentes sobrantes
        bg_trans = PILImage.new("RGBA", img.size, (0, 0, 0, 0))
        diff_trans = ImageChops.difference(img, bg_trans)
        diff_trans_gray = diff_trans.convert("L")
        bbox_trans = diff_trans_gray.getbbox()
        if bbox_trans:
            img = img.crop(bbox_trans)
            
        ext = os.path.splitext(output_path)[1].lower()
        if ext == ".webp":
            img.save(output_path, "WEBP", quality=80)
        else:
            img.save(output_path, "PNG")
        return output_path
    except Exception as e:
        print(f"  [AVISO] No se pudo generar logo blanco de importadora: {e}")
def normalizar_codigo(cod):
    if cod is None:
        return ""
    # Convertir a texto y limpiar espacios normales e invisibles (\xa0, \u200b, \ufeff)
    s = str(cod).replace("\xa0", " ").replace("\u200b", "").replace("\ufeff", "").strip()
    # Si viene como flotante de Excel tipo "1024.0", dejarlo en "1024"
    if re.match(r'^\d+\.0$', s):
        s = s[:-2]
    return s

def clave_busqueda(cod):
    # Clave de búsqueda limpia: sólo alfanuméricos en mayúsculas
    s = normalizar_codigo(cod)
    return re.sub(r'[^A-Za-z0-9]', '', s).upper()

def buscar_producto_en_db(cod, db, db_norm=None, db_clean=None):
    if not cod:
        return None
    raw = normalizar_codigo(cod)
    if raw in db:
        return db[raw]
    if db_norm is not None:
        norm = raw.upper()
        if norm in db_norm:
            return db_norm[norm]
    if db_clean is not None:
        clean = clave_busqueda(raw)
        if clean in db_clean:
            return db_clean[clean]
    return None

def detectar_hojas_inventario(wb):
    """
    Identifica las hojas que contienen inventario/productos.
    Prioriza 'FORMATO INVENTARIO' o nombres similares, o todas las hojas excepto vistas/catálogos.
    """
    candidatas = []
    # 1. Coincidencia por nombre de inventario o formato
    for name in wb.sheetnames:
        name_clean = name.strip().upper()
        if "FORMATO" in name_clean or "INVENTARIO" in name_clean or "INVENT" in name_clean:
            candidatas.append(wb[name])
            
    if candidatas:
        return candidatas
        
    # 2. Si no hay con esas palabras, tomar todas excepto las de vista o catálogo generado
    for name in wb.sheetnames:
        name_clean = name.strip().upper()
        if not any(k in name_clean for k in ["VISTA", "CATALOGO", "CONFIG", "RESUMEN", "PORTADA"]):
            candidatas.append(wb[name])
            
    if not candidatas:
        candidatas = [wb.active or wb[wb.sheetnames[0]]]
        
    return candidatas

def detectar_columnas(ws):
    """
    Detecta automáticamente las posiciones de las columnas de producto leyendo las primeras filas.
    Retorna diccionario con índices (1-based) y la fila de inicio de datos.
    """
    cols = {
        "categoria": COL_CATEGORIA,
        "codigo": COL_CODIGO,
        "imagen": COL_IMAGEN,
        "tipo": COL_TIPO,
        "nombre": COL_NOMBRE,
        "size": COL_SIZE,
        "detalle": COL_DETALLE,
        "uni": COL_UNI,
        "cant_caja": None,
    }
    start_row = FILA_INICIO_DB
    
    max_scan_header = min(10, ws.max_row) if ws.max_row else 10
    found_headers = False
    
    for r in range(1, max_scan_header + 1):
        row_vals = [str(ws.cell(row=r, column=c).value or "").strip().upper() for c in range(1, 20)]
        for idx, val in enumerate(row_vals, start=1):
            if not val:
                continue
            if any(k in val for k in ["CODIGO", "COD.", "ITEM", "REFERENCIA"]) and "BARRAS" not in val:
                cols["codigo"] = idx
                found_headers = True
            elif any(k in val for k in ["CATEGORIA", "RUBRO", "FAMILIA"]):
                cols["categoria"] = idx
            elif any(k in val for k in ["TIPO", "MARCA", "LINEA"]):
                cols["tipo"] = idx
            elif any(k in val for k in ["NOMBRE", "DESCRIPCION", "DESCRIP", "PRODUCTO"]) and "DETALLE" not in val:
                cols["nombre"] = idx
            elif any(k in val for k in ["MEDIDA", "SIZE", "TAMANO", "TAMAÑO", "DIMENSION"]):
                cols["size"] = idx
            elif any(k in val for k in ["DETALLE", "ESPECIFICACION", "CARACTERISTICA", "OBSERVACION"]):
                cols["detalle"] = idx
            elif any(k in val for k in ["POR CAJA", "Q. POR CAJA", "Q.POR CAJA", "Q. POR", "CANT. CAJA", "CANT/CAJA", "X CAJA"]):
                cols["cant_caja"] = idx
            elif any(k in val for k in ["UNI", "UNIDAD", "U.M.", "EMPAQUE", "PRESENTACION"]):
                cols["uni"] = idx
            elif any(k in val for k in ["IMAGEN", "FOTO", "IMG"]):
                cols["imagen"] = idx
                
        if found_headers:
            start_row = r + 1
            break
            
    return cols, start_row

def extraer_imagenes_db(ws_db, filas_interes=None):
    """
    Extrae las imágenes de la hoja de base de datos directamente de los objetos en memoria.
    Devuelve dict: { fila: bytes_imagen }
    Si se pasa filas_interes, solo extrae las imágenes que estén en esas filas específicas.
    """
    imagenes_por_fila = {}
    if not hasattr(ws_db, '_images'):
        return imagenes_por_fila
    print(f"  Imágenes detectadas en hoja '{ws_db.title}': {len(ws_db._images)}")
    for img in ws_db._images:
        try:
            ancla = img.anchor
            if hasattr(ancla, '_from'):
                fila = ancla._from.row + 1
            elif hasattr(ancla, 'row'):
                fila = ancla.row + 1
            else:
                continue
            
            # OPTIMIZACIÓN: Omitir la extracción de bytes si la fila no nos interesa
            if filas_interes is not None and fila not in filas_interes:
                continue
                
            if hasattr(img, 'ref') and img.ref:
                img.ref.seek(0)
                img_bytes = img.ref.read()
                img.ref.seek(0)
                imagenes_por_fila[fila] = img_bytes
        except Exception as e:
            pass
    return imagenes_por_fila

def get_color(color_str, default="000000"):
    """
    Formatea un color hex o palabra clave de color para CSS.
    """
    if not color_str:
        return f"#{default}"
    color_str = str(color_str).strip()
    if color_str.startswith("#"):
        return color_str
    if re.match(r'^[a-fA-F0-9]{3,8}$', color_str):
        return f"#{color_str}"
    return color_str

def exportar_a_pdf_edge(html_path, pdf_path):
    """
    Exporta un archivo HTML a PDF utilizando Microsoft Edge headless.
    """
    edge_paths = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    ]
    
    edge_executable = None
    for path in edge_paths:
        if os.path.exists(path):
            edge_executable = path
            break
            
    if not edge_executable:
        print("\n[ERROR] No se encontró Microsoft Edge o Google Chrome en las rutas estándar.")
        return False
        
    abs_html = os.path.abspath(html_path)
    abs_pdf = os.path.abspath(pdf_path)
    
    if os.path.exists(abs_pdf):
        try:
            os.remove(abs_pdf)
        except Exception as e:
            print(f"[ERROR] No se pudo eliminar el archivo PDF existente: {e}")
            return False
            
    print(f"  Generando PDF con: {edge_executable}")
    cmd = [
        edge_executable,
        "--headless",
        "--no-header-footer",
        f"--print-to-pdf={abs_pdf}",
        abs_html
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if os.path.exists(abs_pdf):
            print(f"  [OK] PDF generado con éxito en: {abs_pdf}")
            return True
        else:
            print(f"  [ERROR] El navegador no generó el PDF. Detalle: {result.stdout} {result.stderr}")
            return False
    except Exception as e:
        print(f"  [ERROR] Ocurrió un error al convertir HTML a PDF: {e}")
        return False

def exportar_a_pdf(excel_path, pdf_path, sheet_name, forzar_imagenes=False):
    """
    Método para compatibilidad con el backend web.
    Redirige a la generación de HTML y conversión a PDF.
    """
    print("\n>>> Iniciando exportación rápida a PDF desde base de datos...")
    return generar(descargar_nube=False, forzar_imagenes=forzar_imagenes)

def generar_html_y_imagenes(db, codigos, imagenes_por_fila, layout="desktop", output_filename="catalogos.html", forzar_imagenes=False, db_norm=None, db_clean=None, whatsapp_phone=None):
    import os
    import re
    import urllib.parse
    def to_base64_src(path_or_bytes, default_mime="image/png"):
        if not path_or_bytes:
            return ""
        try:
            if isinstance(path_or_bytes, bytes):
                import base64
                encoded = base64.b64encode(path_or_bytes).decode('utf-8')
                return f"data:{default_mime};base64,{encoded}"
            elif isinstance(path_or_bytes, str) and os.path.exists(path_or_bytes):
                import base64
                ext = os.path.splitext(path_or_bytes)[1].lower()
                mime = "image/png"
                if ext == ".webp":
                    mime = "image/webp"
                elif ext in (".jpg", ".jpeg"):
                    mime = "image/jpeg"
                with open(path_or_bytes, "rb") as f:
                    encoded = base64.b64encode(f.read()).decode('utf-8')
                return f"data:{mime};base64,{encoded}"
        except Exception as e:
            print(f"  [AVISO] No se pudo codificar imagen a base64: {e}")
        return path_or_bytes
    temp_dir = "temp_imgs"
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
        
    # Cargar base de hashes para invalidación de caché de imágenes de productos
    hashes_path = os.path.join(temp_dir, "image_hashes.json")
    image_hashes = {}
    if not forzar_imagenes and os.path.exists(hashes_path):
        try:
            with open(hashes_path, "r", encoding="utf-8") as f_h:
                image_hashes = json.load(f_h)
        except Exception:
            pass
    if forzar_imagenes:
        print("  [INFO] Forzando regeneración de todas las imágenes. Ignorando caché previa...")
    hashes_updated = forzar_imagenes
        
    # Variables de diseño condicional según el layout (Desktop A4 vs Celular)
    page_size_css = """
    @page {
      size: A4 portrait;
      margin: 12mm 10mm 12mm 10mm;
    }
    """ if layout != "mobile" else """
    @page {
      size: 108mm 192mm;
      margin: 8mm 6mm 8mm 6mm;
    }
    """
    
    grid_cols_css = "grid-template-columns: repeat(3, 1fr) !important;"
    min_height_css = "min-height: 272mm;" if layout != "mobile" else "min-height: 176mm;"
    cover_title_size = "28pt"
    brand_banner_height = "70px"
    brand_logo_height = "55px"
    
    # ─── GENERACIÓN Y RECORTE DE LOGO BLANCO DE LA EMPRESA ───
    impor_logo_cropped = "Logo Impor.png"
    if os.path.exists("Logo Impor.png"):
        impor_logo_cropped = os.path.join(temp_dir, "logo_impor_blanco.webp")
        generar_logo_blanco_empresa("Logo Impor.png", impor_logo_cropped)
        
    # Organizar estructura
    brands_prods = {}       # { brand: { category: [products] } }
    brands_orden = []       # Orden de las marcas
    brands_cats_orden = {}  # { brand: [categories] }
    no_encontrados = []
    
    if db_norm is None:
        db_norm = {normalizar_codigo(k).upper(): v for k, v in db.items()}
    if db_clean is None:
        db_clean = {clave_busqueda(k): v for k, v in db.items()}

    rows_with_products = {p["fila_db"] for p in db.values()}
    
    print("\nProcesando y guardando imágenes de productos...")
    for cod in codigos:
        prod = buscar_producto_en_db(cod, db, db_norm, db_clean)
        if not prod:
            no_encontrados.append(cod)
            continue
        
        brand_raw = prod.get("tipo", "OTRO") or "OTRO"
        brand_theme = get_brand_theme(brand_raw)
        brand_name = brand_theme["display_name"]
        
        cat = prod["categoria"] or "Sin Categoría"
        
        if brand_name not in brands_prods:
            brands_prods[brand_name] = {}
            brands_orden.append(brand_name)
            brands_cats_orden[brand_name] = []
            
        if cat not in brands_prods[brand_name]:
            brands_prods[brand_name][cat] = []
            brands_cats_orden[brand_name].append(cat)
            
        # Extraer imagen
        fila_db = prod["fila_db"]
        ws_title = prod.get("ws_title", "")
        img_bytes = imagenes_por_fila.get((ws_title, fila_db)) or imagenes_por_fila.get(fila_db)
        if not img_bytes:
            img_bytes = imagenes_por_fila.get((ws_title, fila_db - 1)) or imagenes_por_fila.get(fila_db - 1)
            if not img_bytes:
                img_bytes = imagenes_por_fila.get((ws_title, fila_db + 1)) or imagenes_por_fila.get(fila_db + 1)
                
        img_relative_path = ""
        if img_bytes:
            try:
                clean_cod = re.sub(r'[\\/*?:"<>| ]', "_", prod["cod"])
                img_filename = f"prod_{clean_cod}.webp"
                img_path = os.path.join(temp_dir, img_filename)
                
                # Calcular el hash MD5 de los bytes de imagen originales del Excel
                img_hash = hashlib.md5(img_bytes).hexdigest()
                
                # Si el archivo no existe o el hash cambió, procesamos e invalidamos la caché
                if not os.path.exists(img_path) or image_hashes.get(clean_cod) != img_hash:
                    with open(img_path, "wb") as f_img:
                        f_img.write(img_bytes)
                    
                    # Auto recortar márgenes en blanco de la imagen y optimizar tamaño
                    try:
                        img_pil = PILImage.open(img_path)
                        img_cropped = autocrop_image(img_pil)
                        
                        # Redimensionar si es muy grande para reducir peso manteniendo calidad
                        max_dim = 400
                        if img_cropped.width > max_dim or img_cropped.height > max_dim:
                            resample_filter = getattr(PILImage, "Resampling", None)
                            filter_type = resample_filter.LANCZOS if resample_filter else getattr(PILImage, "ANTIALIAS", 3)
                            img_cropped.thumbnail((max_dim, max_dim), filter_type)
                            
                        img_cropped.save(img_path, "WEBP", quality=60)
                    except Exception:
                        pass
                        
                    image_hashes[clean_cod] = img_hash
                    hashes_updated = True
                
                img_relative_path = f"{temp_dir}/{img_filename}"
            except Exception as e:
                print(f"  [AVISO] No se pudo guardar imagen para {prod['cod']}: {e}")
                
        prod_copy = prod.copy()
        prod_copy["img_path"] = img_relative_path
        brands_prods[brand_name][cat].append(prod_copy)
        
    total_prods = sum(len(brands_prods[b][c]) for b in brands_prods for c in brands_prods[b])
    total_marcas = len(brands_orden)
    
    # Función auxiliar para determinar la plantilla de tarjeta según la marca
    def get_card_layout_class(brand_name_str):
        bname = str(brand_name_str).upper()
        if "UYUSTOOLS" in bname:
            return "card-bold"
        elif "FERRAWYY" in bname or "GATE" in bname:
            return "card-vibrant"
        elif any(x in bname for x in ["FERTON", "OMEGA"]):
            return "card-dark-luxury"
        else:
            return "card-tech"
            
    # Función auxiliar para saber si es tema oscuro
    def is_dark_theme(brand_name_str):
        bname = str(brand_name_str).upper()
        return any(x in bname for x in ["FERTON", "OMEGA"])
    
    # ─── CONSTRUIR HTML ───
    html_out = []
    
    html_template = """<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=1200">
  <title>Catálogo de Productos - Importadora Rivero</title>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=Inter:wght@400;600;700;800&display=swap');
    
    html {
      min-width: 1200px;
      width: 1200px;
      margin: 0 auto;
      background-color: #F1F5F9;
    }

    /* PLACEHOLDER_PAGE_SIZE */
    body {
      font-family: 'Plus Jakarta Sans', 'Inter', system-ui, sans-serif;
      background-color: #F1F5F9;
      color: #0F172A;
      margin: 0 auto;
      padding: 0;
      width: 1200px;
      min-width: 1200px;
      max-width: 1200px;
      box-sizing: border-box;
      -webkit-print-color-adjust: exact;
      print-color-adjust: exact;
      /* PLACEHOLDER_BODY_STYLE */
    }
    
    /* Portada Estilo Revista con Texturas y Formas (Oscura como la Foto 1) */
    .cover-page {
      background: radial-gradient(circle at 80% 20%, #1E2530 0%, #0B0E14 100%);
      border-radius: 20px;
      padding: 30px 45px 25px 45px;
      box-sizing: border-box;
      min-height: auto;
      display: flex;
      flex-direction: column;
      justify-content: flex-start;
      page-break-after: always;
      break-after: page;
      position: relative;
      overflow: hidden;
      box-shadow: 0 10px 40px rgba(0,0,0,0.03);
    }
    /* Doble marco de lujo translúcido */
    .cover-page::before {
      content: "";
      position: absolute;
      top: 18px;
      left: 18px;
      right: 18px;
      bottom: 18px;
      border: 1px solid rgba(255, 255, 255, 0.08);
      border-radius: 16px;
      pointer-events: none;
      z-index: 5;
    }
    .cover-page::after {
      content: "";
      position: absolute;
      top: 23px;
      left: 23px;
      right: 23px;
      bottom: 23px;
      border: 1px solid rgba(255, 255, 255, 0.03);
      border-radius: 14px;
      pointer-events: none;
      z-index: 5;
    }
    
    /* Formas decorativas desenfocadas de fondo */
    .decor-circle {
      position: absolute;
      border-radius: 50%;
      filter: blur(90px);
      z-index: 1;
      opacity: 0.15;
    }
    .decor-1 {
      width: 400px;
      height: 400px;
      background: #F97316; /* Naranja */
      top: -150px;
      right: -100px;
    }
    .decor-2 {
      width: 500px;
      height: 500px;
      background: #005BAC; /* Azul */
      bottom: -200px;
      left: -150px;
    }
    
    /* Agrupado hacia arriba para reducir espacios vacíos innecesarios */
    .cover-content {
      position: relative;
      z-index: 10;
      display: flex;
      flex-direction: column;
      justify-content: flex-start;
      height: 100%;
      flex-grow: 1;
    }
    
    /* Cabecera ultra-ajustada arriba y abajo (lo más corta posible) */
    .cover-header {
      display: flex;
      justify-content: space-between;
      align-items: center;
      width: 100%;
      margin-top: 0px;
      margin-bottom: 0px;
      padding: 0;
    }
    .cover-logo-wrapper {
      margin: 14px 0 16px 0;
      display: flex;
      align-items: center;
    }
    .company-logo {
      height: 55px;
      max-height: 60px;
      max-width: 260px;
      object-fit: contain;
      filter: drop-shadow(0 4px 10px rgba(0, 0, 0, 0.4));
    }
    .company-name-fallback {
      font-size: 18pt;
      font-weight: 800;
      color: #FFFFFF;
      margin: 0;
      letter-spacing: 1px;
    }
    .header-date-box {
      border: 1px solid rgba(255, 255, 255, 0.25);
      padding: 4px 10px;
      font-size: 8.5pt;
      font-weight: 700;
      letter-spacing: 2px;
      text-transform: uppercase;
      color: rgba(255, 255, 255, 0.85);
      border-radius: 2px;
    }
    
    /* Traemos el contenido alineado a la misma altura superior */
    .cover-main-content {
      margin-top: 10px;
      margin-bottom: 15px;
      text-align: left;
      padding-left: 0px;
      z-index: 10;
      position: relative;
    }
    .cover-label {
      font-size: 10pt;
      font-weight: 800;
      color: #F97316; /* Naranja */
      letter-spacing: 3px;
      text-transform: uppercase;
      margin: 0;
    }
    .cover-main-title {
      font-size: /* PLACEHOLDER_COVER_TITLE_SIZE */;
      font-weight: 800;
      line-height: 1.15;
      color: #FFFFFF;
      margin: 0 0 15px 0;
      letter-spacing: 1px;
      text-transform: uppercase;
    }
    .title-highlight {
      color: #F97316; /* Naranja */
    }
    .cover-description {
      font-size: 10.5pt;
      line-height: 1.6;
      color: #94A3B8; /* Gris azulado suave */
      max-width: 580px;
      margin: 0;
    }
    /* Cuadraditos elegantes de logos de marcas en la portada */
    .cover-brands-strip {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-top: 18px;
      align-items: center;
    }
    .cover-brand-mini-card {
      border-radius: 8px;
      padding: 4px 12px;
      height: 32px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      cursor: pointer;
      transition: all 0.2s ease;
      box-sizing: border-box;
      overflow: hidden;
    }
    .cover-brand-mini-card:hover {
      transform: translateY(-2px) scale(1.08);
      filter: brightness(1.08);
    }
    .cover-brand-mini-card img {
      max-height: 22px;
      max-width: 80px;
      object-fit: contain;
    }
    .cover-brand-mini-fallback {
      font-size: 8pt;
      font-weight: 800;
      color: #0F172A;
      letter-spacing: 0.5px;
    }
    
    /* El pie de página se empuja al final de forma natural mediante margin-top: auto */
    .cover-footer {
      width: 100%;
      display: flex;
      justify-content: flex-start;
      margin-top: auto;
      margin-bottom: 5px;
      padding-left: 20px;
    }
    .cover-footer-text {
      font-size: 9pt;
      color: rgba(255, 255, 255, 0.35);
      letter-spacing: 1px;
      text-transform: uppercase;
    }
    
    /* Secciones de Marca y Producto con Fondos Adaptados */
    .brand-section {
      page-break-before: always;
      break-before: page;
      padding: 15px 25px;
      margin-bottom: 0;
      /* PLACEHOLDER_MIN_HEIGHT */
      box-sizing: border-box;
    }
    
    /* Encabezado de marcas ajustado para que no sea espacioso */
    .brand-banner {
      width: 100%;
      height: /* PLACEHOLDER_BRAND_BANNER_HEIGHT */;
      display: flex;
      justify-content: center;
      align-items: center;
      padding: 6px 15px;
      border-radius: 12px;
      margin-bottom: 5px;
      box-sizing: border-box;
      page-break-after: avoid;
      break-after: avoid;
      box-shadow: 0 4px 15px rgba(0,0,0,0.02);
    }
    .brand-logo {
      max-height: /* PLACEHOLDER_BRAND_LOGO_HEIGHT */;
      max-width: 95%;
      object-fit: contain;
    }
    .brand-logo-fallback {
      font-size: 20pt;
      font-weight: 800;
      text-transform: uppercase;
      letter-spacing: 1px;
    }
    .brand-subtitle {
      text-align: center;
      font-size: 10pt;
      font-weight: 800;
      margin-bottom: 15px;
      text-transform: uppercase;
      letter-spacing: 1.5px;
      page-break-after: avoid;
      break-after: avoid;
    }
    .category-section {
      margin-bottom: 25px;
    }
    .category-header {
      font-size: 11pt;
      font-weight: 800;
      padding: 8px 14px;
      border-radius: 8px;
      margin: 15px 0 10px 0;
      text-transform: uppercase;
      letter-spacing: 0.5px;
      box-shadow: 0 2px 8px rgba(0,0,0,0.01);
      page-break-after: avoid;
      break-after: avoid;
    }
    .products-grid {
      display: grid;
      /* PLACEHOLDER_GRID_COLS */
      gap: 18px;
      margin-bottom: 20px;
    }

    .product-card {
      position: relative;
      transition: transform 0.2s ease, opacity 0.3s ease, border-color 0.2s ease;
    }
    .btn-card-remove-live {
      position: absolute;
      top: 6px;
      right: 6px;
      width: 22px;
      height: 22px;
      border-radius: 50%;
      background: #EF4444;
      color: #FFFFFF;
      border: 2px solid #FFFFFF;
      font-size: 11px;
      font-weight: 800;
      cursor: pointer;
      display: none !important;
      align-items: center;
      justify-content: center;
      z-index: 50;
      box-shadow: 0 3px 8px rgba(0, 0, 0, 0.4);
      transition: all 0.2s ease;
      line-height: 1;
      padding: 0;
    }
    /* El botón de eliminar SOLO es visible dentro del panel del generador (administrador) */
    body.is-generator-iframe .product-card:hover .btn-card-remove-live {
      display: flex !important;
    }
    .btn-card-remove-live:hover {
      background: #B91C1C;
      transform: scale(1.25);
    }
    
    /* ──── 1. TECH CARD (DongCheng, etc.) ──── */
    .card-tech {
      background: #FFFFFF;
      border: 1px solid rgba(15, 23, 42, 0.08);
      border-radius: 14px;
      overflow: hidden;
      display: flex;
      flex-direction: column;
      box-shadow: 0 4px 15px rgba(0, 0, 0, 0.02);
      border-top: 4px solid var(--brand-header);
      page-break-inside: avoid;
      break-inside: avoid;
    }
    .card-tech .card-header {
      background: rgba(15, 23, 42, 0.01);
      color: var(--brand-header);
      padding: 7px 9px 3px 9px;
      font-size: 9pt;
      font-weight: 800;
      text-align: left;
    }
    .card-tech .card-body {
      padding: 9px 8px;
      display: flex;
      flex-direction: column;
      flex-grow: 1;
    }
    .card-tech .product-name {
      font-size: 9.5pt;
      font-weight: 800;
      color: #1E293B;
      margin: 0 0 8px 0;
      line-height: 1.25;
      height: 2.5em;
      overflow: hidden;
      display: -webkit-box;
      -webkit-line-clamp: 2;
      -webkit-box-orient: vertical;
    }
    .card-tech .image-container {
      background: #FFFFFF;
      height: 135px;
      display: flex;
      justify-content: center;
      align-items: center;
      margin-bottom: 8px;
      padding: 6px;
      border-radius: 8px;
      border: 1px solid rgba(15, 23, 42, 0.03);
    }
    .card-tech .product-img {
      max-width: 100%;
      max-height: 100%;
      object-fit: contain;
    }
    .card-tech .measure-pill {
      text-align: center;
      font-size: 8.5pt;
      font-weight: 800;
      padding: 4px 8px;
      border-radius: 20px;
      margin-bottom: 8px;
      background: var(--brand-measure-bg);
      color: var(--brand-measure-fg);
      border: 1px solid rgba(15, 23, 42, 0.03);
    }
    .card-tech .card-footer {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-top: auto;
      padding-top: 6px;
      border-top: 1px dashed rgba(15, 23, 42, 0.08);
      gap: 3px;
    }
    .card-tech .packaging-info {
      font-size: 8pt;
      font-weight: 800;
      color: #64748B;
    }
    .card-tech .availability-pill {
      background: #DCFCE7;
      color: #15803D;
      font-size: 8pt;
      font-weight: 800;
      padding: 3px 8px;
      border-radius: 12px;
    }
    
    /* ──── 2. BOLD CARD (Uyustools) ──── */
    .card-bold {
      background: #FFFFFF;
      border: 2px solid #0F172A;
      border-radius: 14px;
      overflow: hidden;
      display: flex;
      flex-direction: column;
      box-shadow: 5px 5px 0px #0F172A;
      page-break-inside: avoid;
      break-inside: avoid;
    }
    .card-bold .card-header {
      background: #0F172A;
      color: var(--brand-header);
      padding: 7px 9px;
      font-size: 9pt;
      font-weight: 800;
      text-align: center;
      letter-spacing: 0.5px;
    }
    .card-bold .card-body {
      padding: 9px 8px;
      display: flex;
      flex-direction: column;
      flex-grow: 1;
    }
    .card-bold .product-name {
      font-size: 9.5pt;
      font-weight: 800;
      color: #0F172A;
      margin: 0 0 8px 0;
      line-height: 1.25;
      height: 2.5em;
      overflow: hidden;
      display: -webkit-box;
      -webkit-line-clamp: 2;
      -webkit-box-orient: vertical;
      text-transform: uppercase;
    }
    .card-bold .image-container {
      background: #FFFFFF;
      height: 135px;
      display: flex;
      justify-content: center;
      align-items: center;
      margin-bottom: 8px;
      padding: 6px;
      border: 2px solid #0F172A;
      border-radius: 8px;
    }
    .card-bold .product-img {
      max-width: 100%;
      max-height: 100%;
      object-fit: contain;
    }
    .card-bold .measure-pill {
      text-align: center;
      font-size: 8.5pt;
      font-weight: 800;
      padding: 4px 8px;
      border-radius: 6px;
      margin-bottom: 8px;
      background: var(--brand-header);
      color: #0F172A;
      border: 2px solid #0F172A;
    }
    .card-bold .card-footer {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-top: auto;
      padding-top: 6px;
      border-top: 2px solid #0F172A;
      gap: 3px;
    }
    .card-bold .packaging-info {
      font-size: 7.5pt;
      font-weight: 700;
      color: #0F172A;
    }
    .card-bold .availability-pill {
      background: #0F172A;
      color: var(--brand-header);
      font-size: 7.5pt;
      font-weight: 800;
      padding: 3px 8px;
      border-radius: 6px;
    }
    
    /* ──── 3. VIBRANT CARD (Ferrawyy, etc.) ──── */
    .card-vibrant {
      background: linear-gradient(180deg, #FFFFFF 0%, #FFFDFB 100%);
      border: 1px solid var(--brand-measure-bg);
      border-radius: 18px;
      overflow: hidden;
      display: flex;
      flex-direction: column;
      box-shadow: 0 8px 25px rgba(249, 115, 22, 0.06);
      page-break-inside: avoid;
      break-inside: avoid;
    }
    .card-vibrant .card-header {
      background: linear-gradient(135deg, var(--brand-header) 0%, var(--brand-measure-fg) 100%);
      color: #FFFFFF;
      padding: 7px 9px;
      font-size: 9pt;
      font-weight: 800;
      text-align: center;
      border-bottom-left-radius: 10px;
      border-bottom-right-radius: 10px;
      margin: 0 8px;
      box-shadow: 0 4px 10px rgba(249, 115, 22, 0.15);
    }
    .card-vibrant .card-body {
      padding: 9px 8px;
      display: flex;
      flex-direction: column;
      flex-grow: 1;
    }
    .card-vibrant .product-name {
      font-size: 9.5pt;
      font-weight: 800;
      color: #2D3748;
      margin: 0 0 8px 0;
      line-height: 1.25;
      height: 2.5em;
      overflow: hidden;
      display: -webkit-box;
      -webkit-line-clamp: 2;
      -webkit-box-orient: vertical;
    }
    .card-vibrant .image-container {
      background: #FFFFFF;
      height: 135px;
      display: flex;
      justify-content: center;
      align-items: center;
      margin-bottom: 8px;
      padding: 6px;
      border-radius: 10px;
      border: 1px solid rgba(249, 115, 22, 0.08);
    }
    .card-vibrant .product-img {
      max-width: 100%;
      max-height: 100%;
      object-fit: contain;
    }
    .card-vibrant .measure-pill {
      text-align: center;
      font-size: 8.5pt;
      font-weight: 800;
      padding: 4px 8px;
      border-radius: 20px;
      margin-bottom: 8px;
      background: var(--brand-measure-bg);
      color: var(--brand-measure-fg);
    }
    .card-vibrant .card-footer {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-top: auto;
      padding-top: 6px;
      border-top: 1px dashed var(--brand-measure-bg);
      gap: 3px;
    }
    .card-vibrant .packaging-info {
      font-size: 8pt;
      font-weight: 800;
      color: var(--brand-measure-fg);
    }
    .card-vibrant .availability-pill {
      background: #FFEDD5;
      color: var(--brand-measure-fg);
      font-size: 8pt;
      font-weight: 800;
      padding: 3px 8px;
      border-radius: 10px;
    }
    
    /* ──── 4. LUXURY DARK CARD (Ferton, Omega, Rio, Crown) ──── */
    .card-dark-luxury {
      background: #1E293B;
      border: 1px solid rgba(255, 255, 255, 0.06);
      border-radius: 16px;
      overflow: hidden;
      display: flex;
      flex-direction: column;
      box-shadow: 0 12px 25px rgba(0, 0, 0, 0.35);
      border-top: 4px solid var(--brand-header);
      page-break-inside: avoid;
      break-inside: avoid;
    }
    .card-dark-luxury .card-header {
      background: rgba(255, 255, 255, 0.02);
      color: var(--brand-header);
      padding: 7px 9px 3px 9px;
      font-size: 9pt;
      font-weight: 800;
      text-align: left;
    }
    .card-dark-luxury .card-body {
      padding: 9px 8px;
      display: flex;
      flex-direction: column;
      flex-grow: 1;
    }
    .card-dark-luxury .product-name {
      font-size: 9.5pt;
      font-weight: 800;
      color: #FFFFFF;
      margin: 0 0 8px 0;
      line-height: 1.25;
      height: 2.5em;
      overflow: hidden;
      display: -webkit-box;
      -webkit-line-clamp: 2;
      -webkit-box-orient: vertical;
    }
    .card-dark-luxury .image-container {
      background: #FFFFFF;
      height: 135px;
      display: flex;
      justify-content: center;
      align-items: center;
      margin-bottom: 8px;
      padding: 6px;
      border-radius: 8px;
    }
    .card-dark-luxury .product-img {
      max-width: 100%;
      max-height: 100%;
      object-fit: contain;
    }
    .card-dark-luxury .measure-pill {
      text-align: center;
      font-size: 8pt;
      font-weight: 700;
      padding: 4px 8px;
      border-radius: 12px;
      margin-bottom: 8px;
      background: rgba(255, 255, 255, 0.06);
      color: var(--brand-header);
      border: 1px solid rgba(255, 255, 255, 0.08);
    }
    .card-dark-luxury .card-footer {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-top: auto;
      padding-top: 6px;
      border-top: 1px dashed rgba(255, 255, 255, 0.1);
      gap: 3px;
    }
    .card-dark-luxury .packaging-info {
      font-size: 7.5pt;
      font-weight: 600;
      color: #94A3B8;
    }
    .card-dark-luxury .availability-pill {
      background: rgba(16, 185, 129, 0.2);
      color: #34D399;
      font-size: 7.5pt;
      font-weight: 700;
      padding: 3px 8px;
      border-radius: 10px;
    }
    }
    
    /* OMEGA Custom Branding - Black box background with light blue accents */
    .brand-section-omega .card-dark-luxury {
      background: #080C14;
      border: 1px solid rgba(56, 189, 248, 0.25);
      box-shadow: 0 10px 30px rgba(0, 0, 0, 0.55);
    }
    .brand-section-omega .card-dark-luxury .product-name {
      color: #F8FAFC;
    }
    .brand-section-omega .card-dark-luxury .measure-pill {
      background: rgba(56, 189, 248, 0.1);
      color: #38BDF8;
      border: 1px solid rgba(56, 189, 248, 0.25);
    }
    .brand-section-omega .card-dark-luxury .card-footer {
      border-top: 1px dashed rgba(56, 189, 248, 0.15);
    }
    .brand-section-omega .card-dark-luxury .packaging-info {
      color: #94A3B8;
    }
    
    /* RIO Custom Branding - Royal blue border matching packaging */
    .brand-section-rio .card-tech {
      border: 1.5px solid rgba(37, 99, 235, 0.4);
      box-shadow: 0 6px 18px rgba(37, 99, 235, 0.05);
    }
    .brand-section-rio .card-tech .card-header {
      border-bottom: 1.5px solid rgba(37, 99, 235, 0.15);
    }
    
    /* CROWN Custom Branding - Vibrant red border matching brand */
    .brand-section-crown .card-tech {
      border: 1.5px solid #CA0A10;
      border-top: 1.5px solid #CA0A10; /* uniform solid border all around */
      box-shadow: 0 6px 18px rgba(202, 10, 16, 0.05);
    }
    .brand-section-crown .card-tech .card-header {
      background: #CA0A10;
      color: #FFFFFF;
      border-bottom: none;
      padding: 10px 14px;
      font-weight: 700;
    }
    
    /* TOTAL Custom Branding - Teal green border and solid green header matching brand */
    .brand-section-total .card-tech {
      border: 1.5px solid #186E6F;
      border-top: 1.5px solid #186E6F; /* uniform solid border all around */
      box-shadow: 0 6px 18px rgba(24, 110, 111, 0.05);
    }
    .brand-section-total .card-tech .card-header {
      background: #186E6F;
      color: #FFFFFF;
      border-bottom: none;
      padding: 10px 14px;
      font-weight: 700;
    }
    
    /* WADFOW Custom Branding - Blue border and solid blue header matching brand */
    .brand-section-wadfow .card-tech {
      border: 1.5px solid #0A4E9D;
      border-top: 1.5px solid #0A4E9D; /* uniform solid border all around */
      box-shadow: 0 6px 18px rgba(10, 78, 157, 0.05);
    }
    .brand-section-wadfow .card-tech .card-header {
      background: #0A4E9D;
      color: #FFFFFF;
      border-bottom: none;
      padding: 10px 14px;
      font-weight: 700;
    }
    
    /* NEVA Custom Branding - Black border and solid black header with yellow text */
    .brand-section-neva .card-tech {
      border: 1.5px solid #000000;
      border-top: 1.5px solid #000000;
      box-shadow: 0 6px 18px rgba(0, 0, 0, 0.05);
    }
    .brand-section-neva .card-tech .card-header {
      background: #000000;
      color: #FACC15;
      border-bottom: none;
      padding: 10px 14px;
      font-weight: 700;
    }
    
    /* LUTIAN Custom Branding - Lime green border and solid green header matching brand */
    .brand-section-lutian .card-tech {
      border: 1.5px solid #84CC16;
      border-top: 1.5px solid #84CC16;
      box-shadow: 0 6px 18px rgba(132, 204, 22, 0.05);
    }
    .brand-section-lutian .card-tech .card-header {
      background: #84CC16;
      color: #FFFFFF;
      border-bottom: none;
      padding: 10px 14px;
      font-weight: 700;
    }
    
    /* WhatsApp Direct Contact Button */
    .btn-whatsapp {
      display: inline-flex;
      align-items: center;
      gap: 5px;
      background: #25D366;
      color: #FFFFFF !important;
      text-decoration: none;
      font-size: 8.5pt;
      font-weight: 700;
      padding: 4px 10px;
      border-radius: 12px;
      box-shadow: 0 2px 6px rgba(37, 211, 102, 0.3);
      transition: all 0.2s ease;
      cursor: pointer;
    }
    .btn-whatsapp:hover {
      background: #20BA5A;
      transform: translateY(-1px);
      box-shadow: 0 4px 10px rgba(37, 211, 102, 0.45);
    }
    .btn-whatsapp svg {
      width: 12px;
      height: 12px;
      fill: currentColor;
    }

    /* ──── SELECTOR DUAL DE PEDIDOS (CAJAS Y UNIDADES) ──── */
    .card-footer {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-top: auto;
      padding-top: 8px;
      gap: 4px;
    }
    .packaging-info {
      font-size: 8pt;
      font-weight: 800;
      color: #64748B;
      white-space: nowrap;
    }
    .order-selectors-dual {
      display: flex;
      align-items: center;
      gap: 6px;
      margin-left: auto;
    }
    .qty-group {
      display: flex;
      align-items: center;
      gap: 3px;
    }
    .qty-label {
      font-size: 8pt;
      font-weight: 800;
      color: #334155;
      text-transform: uppercase;
      letter-spacing: 0.3px;
    }
    .product-qty-selector {
      display: inline-flex;
      align-items: center;
      background: #F1F5F9;
      border: 1.5px solid #CBD5E1;
      border-radius: 18px;
      padding: 2px 4px;
      gap: 2px;
      transition: all 0.2s ease;
    }
    .product-qty-selector:focus-within {
      border-color: #25D366;
      box-shadow: 0 0 8px rgba(37, 211, 102, 0.4);
      background: #FFFFFF;
    }
    .btn-qty {
      background: #FFFFFF;
      border: 1.5px solid #CBD5E1;
      color: #0F172A;
      width: 21px;
      height: 21px;
      border-radius: 50%;
      display: flex;
      align-items: center;
      justify-content: center;
      font-size: 11pt;
      font-weight: 800;
      cursor: pointer;
      line-height: 1;
      padding: 0;
      transition: all 0.15s ease;
      user-select: none;
    }
    .btn-qty:hover {
      background: #25D366;
      color: #FFFFFF;
      border-color: #25D366;
      transform: scale(1.15);
    }
    .btn-qty:active {
      transform: scale(0.95);
    }
    .input-qty {
      width: 24px;
      border: none;
      background: transparent;
      text-align: center;
      font-family: 'JetBrains Mono', monospace;
      font-size: 8pt;
      font-weight: 800;
      color: #0F172A;
      outline: none;
      -moz-appearance: textfield;
      padding: 0;
      cursor: text;
    }
    .input-qty::-webkit-outer-spin-button,
    .input-qty::-webkit-inner-spin-button {
      -webkit-appearance: none;
      margin: 0;
    }
    .product-card.has-ordered {
      border-color: #25D366 !important;
      box-shadow: 0 6px 20px rgba(37, 211, 102, 0.25) !important;
    }

    /* ──── SEMÁFORO DE STOCK Y EMPAQUE EN TIEMPO REAL ──── */
    .card-header {
      display: flex !important;
      align-items: center !important;
      justify-content: space-between !important;
      gap: 6px !important;
    }
    .stock-status-pill {
      font-size: 7.5pt;
      font-weight: 800;
      padding: 2px 7px;
      border-radius: 12px;
      letter-spacing: 0.2px;
      text-transform: none;
      display: inline-flex;
      align-items: center;
      gap: 4px;
      line-height: 1.2;
      transition: all 0.3s ease;
      white-space: nowrap;
    }
    .stock-checking {
      background: rgba(148, 163, 184, 0.15);
      color: #94A3B8;
      border: 1px solid rgba(148, 163, 184, 0.25);
    }
    .stock-in-stock {
      background: #DCFCE7 !important;
      color: #166534 !important;
      border: 1px solid #86EFAC !important;
      box-shadow: 0 1px 4px rgba(22, 101, 52, 0.12);
    }
    .stock-low {
      background: #FEF3C7 !important;
      color: #B45309 !important;
      border: 1px solid #FCD34D !important;
      box-shadow: 0 1px 6px rgba(180, 83, 9, 0.2);
      animation: pulse-badge 2s infinite ease-in-out;
    }
    @keyframes pulse-badge {
      0%, 100% { transform: scale(1); }
      50% { transform: scale(1.04); }
    }
    .stock-out {
      background: #FEE2E2 !important;
      color: #991B1B !important;
      border: 1px solid #FCA5A5 !important;
    }
    .product-card.is-out-of-stock {
      opacity: 0.82;
      filter: grayscale(0.25);
    }
    .product-card.is-out-of-stock .product-qty-selector {
      opacity: 0.45;
      pointer-events: none;
    }
    .live-stock-indicator {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      background: rgba(255, 255, 255, 0.1);
      border: 1px solid rgba(255, 255, 255, 0.18);
      padding: 4px 10px;
      border-radius: 20px;
      font-size: 8pt;
      font-weight: 700;
      color: #F8FAFC;
      transition: all 0.2s ease;
      cursor: pointer;
      user-select: none;
    }
    .live-stock-indicator:hover {
      background: rgba(255, 255, 255, 0.18);
      border-color: rgba(255, 255, 255, 0.35);
    }
    .pulse-dot-online {
      width: 7px;
      height: 7px;
      background: #22C55E;
      border-radius: 50%;
      box-shadow: 0 0 8px #22C55E;
      animation: pulse-dot 1.5s infinite;
    }
    .pulse-dot-loading {
      width: 7px;
      height: 7px;
      background: #F59E0B;
      border-radius: 50%;
      animation: pulse-dot 0.8s infinite;
    }
    .pulse-dot-cached {
      width: 7px;
      height: 7px;
      background: #94A3B8;
      border-radius: 50%;
    }
    @keyframes pulse-dot {
      0%, 100% { opacity: 1; transform: scale(1); }
      50% { opacity: 0.4; transform: scale(0.85); }
    }
    .stock-toast {
      position: fixed;
      bottom: 85px;
      left: 50%;
      transform: translateX(-50%) translateY(20px);
      background: rgba(15, 23, 42, 0.96);
      color: #FFFFFF;
      backdrop-filter: blur(16px);
      -webkit-backdrop-filter: blur(16px);
      border: 1.5px solid #F59E0B;
      padding: 10px 20px;
      border-radius: 30px;
      font-size: 8.5pt;
      font-weight: 800;
      box-shadow: 0 12px 30px rgba(0, 0, 0, 0.5);
      z-index: 99999;
      opacity: 0;
      pointer-events: none;
      transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
      display: flex;
      align-items: center;
      gap: 8px;
    }
    .stock-toast.show {
      opacity: 1;
      transform: translateX(-50%) translateY(0);
    }
    .input-qty-error {
      border-color: #EF4444 !important;
      animation: shake-input 0.3s ease;
    }
    @keyframes shake-input {
      0%, 100% { transform: translateX(0); }
      25% { transform: translateX(-3px); }
      75% { transform: translateX(3px); }
    }

    /* ──── BARRA DE BÚSQUEDA EN VIVO PARA EL CLIENTE (STICKY FLOTANTE REDONDEADA) ──── */
    .catalog-search-sticky-bar {
      position: sticky;
      top: 12px;
      z-index: 150;
      background: rgba(15, 23, 42, 0.94);
      backdrop-filter: blur(16px);
      -webkit-backdrop-filter: blur(16px);
      border: 1.5px solid rgba(255, 255, 255, 0.16);
      border-radius: 20px;
      padding: 12px 18px;
      margin: 16px auto 24px auto;
      max-width: 1200px;
      width: calc(100% - 32px);
      box-shadow: 0 10px 30px rgba(0, 0, 0, 0.45);
      font-family: 'Plus Jakarta Sans', sans-serif;
      box-sizing: border-box;
    }
    .catalog-search-inner {
      max-width: 1200px;
      margin: 0 auto;
      display: flex;
      align-items: center;
      gap: 10px;
    }
    .search-input-box {
      position: relative;
      flex-grow: 1;
      display: flex;
      align-items: center;
    }
    .search-input-box svg {
      position: absolute;
      left: 12px;
      color: #94A3B8;
      pointer-events: none;
    }
    .search-input-box input {
      width: 100%;
      background: rgba(255, 255, 255, 0.08);
      border: 1.5px solid rgba(255, 255, 255, 0.16);
      border-radius: 25px;
      padding: 9px 34px 9px 36px;
      color: #FFFFFF;
      font-family: inherit;
      font-size: 9pt;
      outline: none;
      transition: all 0.2s ease;
      box-sizing: border-box;
    }
    .search-input-box input:focus {
      background: rgba(255, 255, 255, 0.14);
      border-color: #F59E0B;
      box-shadow: 0 0 12px rgba(245, 158, 11, 0.4);
    }
    .search-input-box input::placeholder {
      color: #94A3B8;
      font-size: 8.5pt;
    }
    .catalog-search-clear-btn {
      position: absolute;
      right: 10px;
      background: rgba(255, 255, 255, 0.16);
      border: none;
      color: #FFFFFF;
      width: 20px;
      height: 20px;
      border-radius: 50%;
      display: none;
      align-items: center;
      justify-content: center;
      cursor: pointer;
      font-size: 8.5pt;
      font-weight: 800;
      transition: all 0.2s;
      line-height: 1;
      padding: 0;
    }
    .catalog-search-clear-btn:hover {
      background: #EF4444;
    }
    .search-stats-badge {
      font-size: 8pt;
      font-weight: 800;
      color: #F59E0B;
      background: rgba(245, 158, 11, 0.15);
      padding: 6px 12px;
      border-radius: 20px;
      white-space: nowrap;
      border: 1px solid rgba(245, 158, 11, 0.3);
      display: flex;
      align-items: center;
      gap: 4px;
    }
    .brand-filter-chips {
      max-width: 1200px;
      margin: 8px auto 0;
      display: flex;
      align-items: center;
      gap: 6px;
      overflow-x: auto;
      padding-bottom: 2px;
    }
    .brand-filter-chips::-webkit-scrollbar {
      height: 3px;
    }
    .brand-filter-chips::-webkit-scrollbar-thumb {
      background: rgba(255, 255, 255, 0.2);
      border-radius: 2px;
    }
    .brand-chip {
      background: rgba(255, 255, 255, 0.06);
      border: 1px solid rgba(255, 255, 255, 0.12);
      color: #CBD5E1;
      padding: 4px 10px;
      border-radius: 14px;
      font-size: 7.5pt;
      font-weight: 700;
      cursor: pointer;
      white-space: nowrap;
      transition: all 0.2s;
    }
    .brand-chip:hover {
      background: rgba(255, 255, 255, 0.15);
      color: #FFFFFF;
    }
    .brand-chip.active {
      background: #F59E0B;
      color: #0F172A;
      border-color: #F59E0B;
      font-weight: 800;
      box-shadow: 0 2px 8px rgba(245, 158, 11, 0.35);
    }
    .no-results-card {
      display: none;
      text-align: center;
      padding: 40px 20px;
      background: #FFFFFF;
      border-radius: 12px;
      margin: 30px auto;
      max-width: 500px;
      border: 1.5px dashed #CBD5E1;
      box-shadow: 0 8px 24px rgba(0,0,0,0.06);
    }
    @media print {
      .catalog-search-sticky-bar, .no-results-card {
        display: none !important;
      }
    }

    /* ──── BARRA FLOTANTE DE COMPRA EN VIVO (PROMINENTE, MODERNA Y FLUIDA) ──── */
    .floating-cart-bar {
      position: fixed;
      bottom: 24px;
      left: 50%;
      transform: translateX(-50%) translateY(160%);
      background: rgba(11, 17, 32, 0.96);
      backdrop-filter: blur(20px);
      -webkit-backdrop-filter: blur(20px);
      border: 1.5px solid rgba(255, 255, 255, 0.18);
      border-radius: 32px;
      padding: 14px 26px;
      display: flex !important;
      align-items: center;
      justify-content: space-between;
      gap: 20px;
      box-shadow: 0 20px 60px rgba(0, 0, 0, 0.75), 0 0 30px rgba(37, 211, 102, 0.25);
      z-index: 9999;
      max-width: 920px;
      width: calc(100% - 60px);
      box-sizing: border-box;
      opacity: 0;
      pointer-events: none;
      transition: transform 0.35s cubic-bezier(0.16, 1, 0.3, 1), opacity 0.3s ease;
      font-family: 'Plus Jakarta Sans', sans-serif;
    }
    .floating-cart-bar.visible {
      transform: translateX(-50%) translateY(0) !important;
      opacity: 1 !important;
      pointer-events: auto !important;
    }
    .cart-summary {
      display: flex;
      align-items: center;
      gap: 14px;
      cursor: pointer;
    }
    .cart-icon-wrapper {
      position: relative;
      background: rgba(37, 211, 102, 0.15);
      border: 1.5px solid rgba(37, 211, 102, 0.35);
      width: 52px;
      height: 52px;
      border-radius: 50%;
      display: flex;
      align-items: center;
      justify-content: center;
      color: #25D366;
    }
    .cart-badge {
      position: absolute;
      top: -4px;
      right: -4px;
      background: #EF4444;
      color: #FFFFFF;
      font-size: 10pt;
      font-weight: 800;
      width: 22px;
      height: 22px;
      border-radius: 50%;
      display: flex;
      align-items: center;
      justify-content: center;
      border: 2px solid #0B1120;
    }
    .cart-text {
      display: flex;
      flex-direction: column;
      gap: 2px;
    }
    .cart-title {
      font-size: 13pt;
      color: #F8FAFC;
      font-weight: 600;
    }
    .cart-title strong {
      color: #25D366;
      font-weight: 800;
    }
    .cart-subtitle {
      font-size: 10pt;
      color: #94A3B8;
    }
    .cart-actions {
      display: flex;
      align-items: center;
      gap: 12px;
    }
    .btn-view-order {
      background: rgba(255, 255, 255, 0.1);
      border: 1px solid rgba(255, 255, 255, 0.18);
      color: #FFFFFF;
      padding: 11px 20px;
      border-radius: 25px;
      font-size: 11.5pt;
      font-weight: 700;
      cursor: pointer;
      display: flex;
      align-items: center;
      gap: 8px;
      font-family: inherit;
      transition: all 0.2s;
      white-space: nowrap;
    }
    .btn-view-order:hover {
      background: rgba(255, 255, 255, 0.18);
      transform: translateY(-1px);
    }
    .btn-send-whatsapp-order {
      background: linear-gradient(135deg, #25D366 0%, #128C7E 100%);
      color: #FFFFFF;
      border: none;
      padding: 12px 24px;
      border-radius: 25px;
      font-size: 12.5pt;
      font-weight: 800;
      cursor: pointer;
      display: flex;
      align-items: center;
      gap: 8px;
      font-family: inherit;
      box-shadow: 0 6px 20px rgba(37, 211, 102, 0.45);
      transition: all 0.2s;
      white-space: nowrap;
    }
    .btn-send-whatsapp-order:hover {
      transform: scale(1.03);
      box-shadow: 0 8px 25px rgba(37, 211, 102, 0.6);
    }

    /* ──── MODAL RESUMEN DE PEDIDO (AMPLIO Y LEGIBLE EN 1200PX) ──── */
    .modal-backdrop {
      position: fixed;
      top: 0; left: 0; right: 0; bottom: 0;
      background: rgba(4, 7, 13, 0.86);
      backdrop-filter: blur(14px);
      -webkit-backdrop-filter: blur(14px);
      display: none;
      justify-content: center;
      align-items: center;
      z-index: 10000;
      padding: 24px;
      font-family: 'Plus Jakarta Sans', sans-serif;
    }
    .modal-backdrop.open {
      display: flex;
    }
    .modal-card {
      background: #0B1120;
      border: 1.5px solid rgba(255, 255, 255, 0.18);
      border-radius: 24px;
      width: 100%;
      max-width: 880px;
      max-height: 90vh;
      display: flex;
      flex-direction: column;
      box-shadow: 0 30px 90px rgba(0, 0, 0, 0.85);
      overflow: hidden;
      color: #FFFFFF;
      animation: modalSlideUp 0.25s ease-out;
    }
    @keyframes modalSlideUp {
      from { transform: translateY(24px); opacity: 0; }
      to { transform: translateY(0); opacity: 1; }
    }
    .modal-header {
      padding: 22px 28px;
      border-bottom: 1px solid rgba(255, 255, 255, 0.1);
      display: flex;
      justify-content: space-between;
      align-items: center;
      background: rgba(255, 255, 255, 0.02);
    }
    .modal-title {
      font-size: 15pt;
      font-weight: 800;
      display: flex;
      align-items: center;
      gap: 10px;
      color: #25D366;
    }
    .modal-close {
      background: rgba(255, 255, 255, 0.08);
      border: 1px solid rgba(255, 255, 255, 0.14);
      color: #CBD5E1;
      font-size: 14pt;
      cursor: pointer;
      width: 36px;
      height: 36px;
      border-radius: 50%;
      display: flex;
      align-items: center;
      justify-content: center;
      transition: all 0.2s;
    }
    .modal-close:hover {
      background: rgba(239, 68, 68, 0.25);
      color: #F87171;
    }
    .modal-body {
      padding: 22px 28px;
      overflow-y: auto;
      display: flex;
      flex-direction: column;
      gap: 16px;
    }
    .customer-inputs {
      display: flex;
      flex-direction: column;
      gap: 10px;
    }
    .customer-inputs-grid {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 10px;
    }
    .customer-field {
      background: #1E293B;
      border: 1.5px solid #475569;
      border-radius: 12px;
      padding: 13px 16px;
      color: #FFFFFF !important;
      font-size: 11pt;
      font-family: inherit;
      font-weight: 600;
      outline: none;
      transition: all 0.2s;
      box-sizing: border-box;
      width: 100%;
    }
    .customer-field::placeholder {
      color: #94A3B8;
      opacity: 1;
    }
    .customer-field:focus {
      border-color: #25D366;
      background: #0F172A;
      box-shadow: 0 0 12px rgba(37, 211, 102, 0.3);
    }
    .order-items-list {
      display: flex;
      flex-direction: column;
      gap: 8px;
      max-height: 280px;
      overflow-y: auto;
    }
    .order-items-list::-webkit-scrollbar {
      width: 6px;
    }
    .order-items-list::-webkit-scrollbar-thumb {
      background: #334155;
      border-radius: 3px;
    }
    .order-item-row {
      display: flex;
      justify-content: space-between;
      align-items: center;
      background: #161F30;
      border: 1px solid rgba(255, 255, 255, 0.08);
      border-radius: 12px;
      padding: 12px 16px;
    }
    .order-item-info {
      display: flex;
      flex-direction: column;
      gap: 3px;
      overflow: hidden;
      padding-right: 12px;
    }
    .order-item-code {
      font-family: 'JetBrains Mono', monospace;
      font-size: 10pt;
      font-weight: 800;
      color: #25D366;
    }
    .order-item-name {
      font-size: 11pt;
      font-weight: 700;
      color: #FFFFFF;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
      max-width: 480px;
    }
    .order-item-meta {
      font-size: 9.5pt;
      color: #94A3B8;
    }
    .order-item-row .product-qty-selector {
      background: #0B1120;
      border-color: #475569;
      padding: 2px 4px;
    }
    .order-item-row .input-qty {
      color: #FFFFFF !important;
      font-weight: 800;
      font-size: 11pt;
      width: 32px;
    }
    .order-item-row .btn-qty {
      background: #334155;
      color: #FFFFFF;
      border-color: #475569;
      width: 26px;
      height: 26px;
      font-size: 12pt;
    }
    .order-item-row .btn-qty:hover {
      background: #25D366;
      color: #0F172A;
      border-color: #25D366;
    }
    .order-total-banner {
      background: rgba(37, 211, 102, 0.12);
      border: 1.5px solid rgba(37, 211, 102, 0.35);
      border-radius: 12px;
      padding: 14px 18px;
      display: flex;
      justify-content: space-between;
      align-items: center;
      font-size: 12pt;
      color: #F8FAFC;
      font-weight: 600;
    }
    .order-total-banner strong {
      font-size: 15pt;
      color: #25D366;
      font-weight: 800;
    }
    .modal-footer {
      padding: 16px 28px;
      border-top: 1px solid rgba(255, 255, 255, 0.1);
      display: flex;
      flex-wrap: wrap;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
      background: rgba(255, 255, 255, 0.02);
    }
    .btn-clear-cart {
      background: transparent;
      border: 1px solid rgba(239, 68, 68, 0.35);
      color: #F87171;
      padding: 10px 18px;
      border-radius: 10px;
      font-size: 10.5pt;
      font-weight: 700;
      cursor: pointer;
      font-family: inherit;
      transition: all 0.2s;
    }
    .btn-clear-cart:hover {
      background: rgba(239, 68, 68, 0.2);
    }
    .btn-copy-sheets {
      background: #1E40AF;
      border: 1px solid #3B82F6;
      color: #FFFFFF;
      padding: 10px 18px;
      border-radius: 10px;
      font-size: 11pt;
      font-weight: 800;
      cursor: pointer;
      font-family: inherit;
      display: flex;
      align-items: center;
      gap: 8px;
      transition: all 0.2s;
      box-shadow: 0 2px 8px rgba(37, 99, 235, 0.3);
    }
    .btn-copy-sheets:hover {
      background: #2563EB;
      box-shadow: 0 4px 12px rgba(37, 99, 235, 0.45);
      transform: translateY(-1px);
    }
    .btn-send-whatsapp-large {
      background: linear-gradient(135deg, #25D366 0%, #128C7E 100%);
      color: #FFFFFF;
      border: none;
      padding: 12px 24px;
      border-radius: 10px;
      font-size: 11.5pt;
      font-weight: 800;
      cursor: pointer;
      display: flex;
      align-items: center;
      gap: 8px;
      font-family: inherit;
      box-shadow: 0 4px 14px rgba(37, 211, 102, 0.4);
      transition: all 0.2s;
    }
    .btn-send-whatsapp-large:hover {
      transform: scale(1.02);
      box-shadow: 0 6px 18px rgba(37, 211, 102, 0.55);
    }

    /* ──── TOAST Y NOTIFICACIONES ──── */
      display: flex;
      align-items: center;
      gap: 6px;
    }

    @media print {
      .products-grid {
        grid-template-columns: 1fr 1fr !important;
        gap: 18px !important;
      }
      .image-container {
        height: 180px !important;
      }
      .product-name {
        font-size: 10.5pt !important;
      }
      .floating-cart-bar,
      .modal-backdrop,
      .welcome-modal-backdrop,
      .product-qty-selector {
        display: none !important;
      }
    }
  </style>
</head>
<body>
"""
    
    body_style_css = """
      max-width: 1200px;
      margin: 0 auto;
      box-shadow: 0 0 30px rgba(0,0,0,0.05);
    """ if layout == "mobile" else ""
    
    html_template_processed = html_template.replace("/* PLACEHOLDER_PAGE_SIZE */", page_size_css)
    html_template_processed = html_template_processed.replace("/* PLACEHOLDER_BODY_STYLE */", body_style_css)
    html_template_processed = html_template_processed.replace("/* PLACEHOLDER_GRID_COLS */", grid_cols_css)
    html_template_processed = html_template_processed.replace("/* PLACEHOLDER_MIN_HEIGHT */", min_height_css)
    html_template_processed = html_template_processed.replace("/* PLACEHOLDER_COVER_TITLE_SIZE */", cover_title_size)
    html_template_processed = html_template_processed.replace("/* PLACEHOLDER_BRAND_BANNER_HEIGHT */", brand_banner_height)
    html_template_processed = html_template_processed.replace("/* PLACEHOLDER_BRAND_LOGO_HEIGHT */", brand_logo_height)
    
    html_out.append(html_template_processed)
    
    # ─── AGREGAR PORTADA ───
    # (Página 1: Portada Limpia estilo Foto 1)
    hoy = date.today()
    mes_nombres = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
                   7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}
    titulo_fecha = f"{mes_nombres[hoy.month]} {hoy.year}"
    
    html_out.append('  <div class="cover-page">')
    # Añadir formas geométricas de fondo para diseño premium
    html_out.append('    <div class="decor-circle decor-1"></div>')
    html_out.append('    <div class="decor-circle decor-2"></div>')
    
    # Contenido flotando sobre el fondo
    html_out.append('    <div class="cover-content">')
    
    # Header con Etiqueta (Izquierda) e Información de Fecha (Derecha) a la misma altura
    html_out.append('      <div class="cover-header">')
    html_out.append('        <div class="cover-label">Catálogo General</div>')
    html_out.append(f'        <div class="header-date-box">{titulo_fecha}</div>')
    html_out.append('      </div>')
    
    # Contenido Principal (Centro-Izquierda)
    html_out.append('      <div class="cover-main-content">')
    html_out.append('        <h1 class="cover-main-title">')
    html_out.append('          Herramientas &<br>Equipos <span class="title-highlight">Profesionales</span>')
    html_out.append('        </h1>')
    
    # Logo de la Empresa más grande y alineado debajo del título
    html_out.append('        <div class="cover-logo-wrapper">')
    if os.path.exists(impor_logo_cropped):
        company_logo_b64 = to_base64_src(impor_logo_cropped)
        html_out.append(f'          <img class="company-logo" src="{company_logo_b64}" alt="Logo Importadora Rivero" />')
    else:
        html_out.append('          <span class="company-name-fallback">IMPORTADORA RIVERO</span>')
    html_out.append('        </div>')
    
    html_out.append('        <p class="cover-description">')
    html_out.append('          Selección de herramientas eléctricas, equipo industrial y accesorios eléctricos, respaldados por marcas de calidad comprobada para el profesional y el distribuidor.')
    html_out.append('        </p>')

    # Cuadraditos interactivos con logos de las marcas (con colores auténticos de cada marca)
    def get_brand_pill_style(brand_name_str):
        bname = str(brand_name_str).upper()
        if "UYUSTOOLS" in bname:
            return "background-color: #FDC800; border: 1.5px solid #EAB308; box-shadow: 0 4px 14px rgba(253, 200, 0, 0.35);"
        elif "FERTON" in bname or "OMEGA" in bname:
            return "background-color: #000000; border: 1.5px solid rgba(251, 191, 36, 0.4); box-shadow: 0 4px 14px rgba(0, 0, 0, 0.6);"
        elif "FERRAWYY" in bname:
            return "background-color: #FFEDD5; border: 1.5px solid #FDBA74; box-shadow: 0 4px 14px rgba(249, 115, 22, 0.25);"
        elif "GATE" in bname:
            return "background-color: #FEE2E2; border: 1.5px solid #FCA5A5; box-shadow: 0 4px 14px rgba(239, 68, 68, 0.25);"
        elif "CROWN" in bname:
            return "background-color: #FFFFFF; border: 1.5px solid #FECACA; box-shadow: 0 4px 14px rgba(202, 10, 16, 0.2);"
        elif "AQUA" in bname:
            return "background-color: #F0F9FF; border: 1.5px solid #BAE6FD; box-shadow: 0 4px 14px rgba(2, 132, 199, 0.2);"
        elif "DONGCHENG" in bname:
            return "background-color: #FFFFFF; border: 1.5px solid #BFDBFE; box-shadow: 0 4px 14px rgba(0, 91, 172, 0.2);"
        elif "NORSTAR" in bname:
            return "background-color: #FFFFFF; border: 1.5px solid #E2E8F0; box-shadow: 0 4px 14px rgba(0, 0, 0, 0.2);"
        elif "DWT" in bname:
            return "background-color: #FFFFFF; border: 1.5px solid #E2E8F0; box-shadow: 0 4px 14px rgba(0, 0, 0, 0.2);"
        else:
            return "background-color: #FFFFFF; border: 1.5px solid rgba(255, 255, 255, 0.35); box-shadow: 0 4px 14px rgba(0, 0, 0, 0.3);"

    html_out.append('        <div class="cover-brands-strip">')
    for b_name_strip in brands_orden:
        b_theme_strip = get_brand_theme(b_name_strip)
        logo_strip_path = b_theme_strip.get("logo")
        safe_b_strip = b_name_strip.replace('"', '&quot;').replace("'", "&#39;")
        clean_strip_bname = re.sub(r'[\\/*?:"<>| ]', "_", b_name_strip)
        dest_strip_logo = os.path.join(temp_dir, f"cropped_logo_{clean_strip_bname}.webp")
        pill_style = get_brand_pill_style(b_name_strip)
        
        html_out.append(f'          <div class="cover-brand-mini-card" style="{pill_style}" onclick="filterByBrand(\'{safe_b_strip}\')" title="Ver productos de {safe_b_strip}">')
        if logo_strip_path and os.path.exists(dest_strip_logo):
            strip_b64 = to_base64_src(dest_strip_logo)
            html_out.append(f'            <img src="{strip_b64}" alt="{safe_b_strip}" />')
        elif logo_strip_path and os.path.exists(logo_strip_path):
            strip_b64 = to_base64_src(logo_strip_path)
            html_out.append(f'            <img src="{strip_b64}" alt="{safe_b_strip}" />')
        else:
            fallback_color = "#FFFFFF" if ("FERTON" in b_name_strip.upper() or "OMEGA" in b_name_strip.upper()) else "#0F172A"
            html_out.append(f'            <span class="cover-brand-mini-fallback" style="color: {fallback_color};">{safe_b_strip}</span>')
        html_out.append('          </div>')
    html_out.append('        </div>')
    html_out.append('      </div>')
    
    # Slogan inferior estilo Foto 1
    html_out.append('      <div class="cover-footer">')
    html_out.append('        <div class="cover-footer-text">')
    html_out.append(f'          • Resumen: {total_prods} productos en {total_marcas} marcas • simplificamos tu esfuerzo •')
    html_out.append('        </div>')
    html_out.append('      </div>')
    
    html_out.append('    </div>') # cover-content
    html_out.append('  </div>') # cover-page
    
    # ─── BARRA DE BÚSQUEDA INTERACTIVA EN VIVO Y FILTRO POR MARCA ───
    html_out.append('  <div id="catalog-search-sticky-bar" class="catalog-search-sticky-bar">')
    html_out.append('    <div class="catalog-search-inner">')
    html_out.append('      <div class="search-input-box">')
    html_out.append('        <svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><circle cx="11" cy="11" r="8"></circle><line x1="21" y1="21" x2="16.65" y2="16.65"></line></svg>')
    html_out.append('        <input type="text" id="catalog-live-search" placeholder="🔍 Buscar por código, nombre, marca o medida (ej: BOM6044, taladro, 8mm)..." oninput="filterCatalogLive(this.value)" autocomplete="off" />')
    html_out.append('        <button type="button" id="catalog-search-clear-btn" class="catalog-search-clear-btn" onclick="clearCatalogSearch()">✕</button>')
    html_out.append('      </div>')
    html_out.append('      <div class="search-stats-badge">')
    html_out.append(f'        <span id="search-match-count">{total_prods}</span> productos')
    html_out.append('      </div>')
    html_out.append('      <div id="live-stock-indicator" class="live-stock-indicator" onclick="fetchLiveStock(true)" title="Clic para refrescar stock en vivo ahora">')
    html_out.append('        <span class="pulse-dot-loading"></span> Sincronizando stock...')
    html_out.append('      </div>')
    html_out.append('    </div>')
    html_out.append('    <div class="brand-filter-chips">')
    html_out.append(f'      <button type="button" class="brand-chip active" data-brand-chip="all" onclick="filterByBrand(\'all\')">Todas ({total_prods})</button>')
    for b_name_chip in brands_orden:
        b_count_chip = sum(len(prods) for prods in brands_prods[b_name_chip].values())
        safe_b_chip = b_name_chip.replace('"', '&quot;').replace("'", "&#39;")
        html_out.append(f'      <button type="button" class="brand-chip" data-brand-chip="{safe_b_chip}" onclick="filterByBrand(\'{safe_b_chip}\')">{safe_b_chip} ({b_count_chip})</button>')
    html_out.append('    </div>')
    html_out.append('  </div>')
    
    # ─── AGREGAR SECCIONES DE MARCAS ───
    for b_name in brands_orden:
        b_theme = get_brand_theme(b_name)
        bg_hdr = get_color(b_theme["header_bg"], "FFFFFF")
        card_hdr_bg = get_color(b_theme["card_header_bg"], "1E293B")
        card_hdr_fg = get_color(b_theme.get("card_header_fg", "FFFFFF"), "FFFFFF")
        card_meas_bg = get_color(b_theme.get("card_measure_bg", "F1F5F9"))
        card_meas_fg = get_color(b_theme.get("card_measure_fg", "0F172A"))
        
        # Determinar estilo de tarjeta y si es tema oscuro
        card_class = get_card_layout_class(b_name)
        is_dark = is_dark_theme(b_name)
        
        # Color de fondo de sección dinámica según branding
        bname_upper = b_name.upper()
        if "UYUSTOOLS" in bname_upper:
            sec_bg = "#FAF7F2"
            fg_sub = "#475569"
        elif "FERRAWYY" in bname_upper or "GATE" in bname_upper:
            sec_bg = "#FFF9F5"
            fg_sub = "#C2410C"
        elif "LUTIAN" in bname_upper:
            sec_bg = "#F7FAF4" # Soft green-tinted light background
            fg_sub = get_color(b_theme["subtitle_color"], "3F6212")
        elif "NEVA" in bname_upper:
            sec_bg = "#FAF9F5" # Warm ivory/linen background
            fg_sub = get_color(b_theme["subtitle_color"], "000000")
        elif is_dark:
            sec_bg = "#0B0F19"
            fg_sub = get_color(b_theme.get("subtitle_color"), "FBBF24")
        else:
            sec_bg = "#F3F7FC"
            fg_sub = get_color(b_theme["subtitle_color"], "0F172A")
            
        clean_brand_class = f"brand-section-{re.sub(r'[^a-zA-Z0-9]', '-', b_name.lower())}"
        html_out.append(f'  <div class="brand-section {clean_brand_class}" data-brand-section="{b_name}" style="background-color: {sec_bg};">')
        
        # Crop del logo para el banner de la marca
        logo_path = b_theme.get("logo")
        logo_src = ""
        if logo_path and os.path.exists(logo_path):
            try:
                clean_bname = re.sub(r'[\\/*?:"<>| ]', "_", b_name)
                dest_logo_path = os.path.join(temp_dir, f"cropped_logo_{clean_bname}.webp")
                if not os.path.exists(dest_logo_path) or os.path.getmtime(logo_path) > os.path.getmtime(dest_logo_path):
                    logo_img = PILImage.open(logo_path)
                    logo_img = autocrop_image(logo_img)
                    if logo_img.width > 300 or logo_img.height > 300:
                        resample_filter = getattr(PILImage, "Resampling", None)
                        filter_type = resample_filter.LANCZOS if resample_filter else getattr(PILImage, "ANTIALIAS", 3)
                        logo_img.thumbnail((300, 300), filter_type)
                    logo_img.save(dest_logo_path, "WEBP", quality=70)
                logo_src = dest_logo_path
            except Exception as e:
                print(f"  [AVISO] No se pudo recortar el logo de la marca {b_name}: {e}")
                logo_src = logo_path

        # Cabecera de marca muy ajustada arriba y abajo
        html_out.append(f'    <div class="brand-banner" style="background-color: {bg_hdr}; border: 1px solid #E2E8F0;">')
        if logo_src:
            brand_logo_b64 = to_base64_src(logo_src)
            html_out.append(f'      <img class="brand-logo" src="{brand_logo_b64}" alt="{b_name}" />')
        else:
            html_out.append(f'      <span class="brand-logo-fallback" style="color: {fg_sub};">{b_name}</span>')
        html_out.append('    </div>')
        
        html_out.append(f'    <div class="brand-subtitle" style="color: {fg_sub};">MARCA: {b_name}</div>')
        
        # Categorías de la Marca
        for cat in brands_cats_orden[b_name]:
            prods = brands_prods[b_name][cat]
            
            if is_dark:
                cat_bg = "rgba(255, 255, 255, 0.05)"
                cat_fg = "#FFFFFF"
            else:
                cat_bg = get_color(b_theme["category_bg"], "F1F5F9")
                cat_fg = get_color(b_theme["category_fg"], "0F172A")
            
            html_out.append(f'    <div class="category-section">')
            # Encabezado de línea de categoría más estrecho
            html_out.append(f'      <div class="category-header" style="background-color: {cat_bg}; color: {cat_fg}; border-left: 4px solid {card_hdr_bg};">')
            html_out.append(f'        • LÍNEA DE {cat.upper()}')
            html_out.append('      </div>')
            
            # Variables de estilo inyectadas en CSS inline para las tarjetas
            html_out.append(f'      <div class="products-grid" style="--brand-header: {card_hdr_bg}; --brand-header-fg: {card_hdr_fg}; --brand-measure-bg: {card_meas_bg}; --brand-measure-fg: {card_meas_fg};">')
            
            for prod in prods:
                search_text = f"{prod['cod']} {prod['nombre']} {prod['size']} {b_name} {cat}".replace('"', '&quot;')
                html_out.append(f'        <div class="{card_class} product-card" data-search="{search_text}" data-brand="{b_name}" data-code="{prod["cod"]}">')
                html_out.append(f'          <button type="button" class="btn-card-remove-live" onclick="quitarProductoEnVivo(event, \'{prod["cod"]}\')" title="Quitar este producto del catálogo">✕</button>')
                html_out.append('          <div class="card-header">')
                html_out.append(f'            <span style="font-weight: 800;">CÓDIGO: {prod["cod"]}</span>')
                html_out.append(f'            <span class="stock-status-pill stock-in-stock" id="stock_pill_{prod["cod"]}">🟢 En Stock</span>')
                html_out.append('          </div>')
                
                html_out.append('          <div class="card-body">')
                html_out.append(f'            <div class="product-name">{prod["nombre"]}</div>')
                
                html_out.append('            <div class="image-container">')
                if prod.get("img_path"):
                    img_b64 = to_base64_src(prod["img_path"])
                    html_out.append(f'              <img class="product-img" src="{img_b64}" alt="{prod["nombre"]}" />')
                else:
                    html_out.append('              <div class="no-photo">SIN FOTO</div>')
                html_out.append('            </div>')
                
                html_out.append('            <div class="measure-pill">')
                html_out.append(f'              Medida: {prod["size"]}')
                html_out.append('            </div>')
                
                # Enlace directo de cotización por WhatsApp
                wa_msg = urllib.parse.quote(f"Hola Importadora Rivero, deseo cotizar el producto:\n* Código: {prod['cod']}\n* Nombre: {prod['nombre']}\n* Marca: {b_name}")
                clean_phone = re.sub(r'[^\d]', '', str(whatsapp_phone or ''))
                safe_name = str(prod["nombre"]).replace('"', '&quot;').replace("'", "&#39;")
                safe_brand = str(b_name).replace('"', '&quot;').replace("'", "&#39;")
                safe_unit = str(prod["uni"]).replace('"', '&quot;').replace("'", "&#39;")

                html_out.append('            <div class="card-footer">')
                html_out.append(f'              <span class="packaging-info" id="pkg_info_{prod["cod"]}">📦 {prod["uni"]}</span>')
                html_out.append('              <div class="order-selectors-dual">')
                html_out.append('                <div class="qty-group" title="Escribe la cantidad de Cajas o usa + / −">')
                html_out.append('                  <span class="qty-label">Caja:</span>')
                html_out.append('                  <div class="product-qty-selector">')
                html_out.append(f'                    <button type="button" class="btn-qty" onclick="stepProductQty(\'{prod["cod"]}\', \'cajas\', -1)">−</button>')
                html_out.append(f'                    <input type="number" id="cajas_{prod["cod"]}" class="input-qty" value="0" min="0" oninput="onDirectInput(\'{prod["cod"]}\')" onchange="onDirectInput(\'{prod["cod"]}\')" data-code="{prod["cod"]}" data-name="{safe_name}" data-brand="{safe_brand}" data-unit="{safe_unit}" placeholder="0" />')
                html_out.append(f'                    <button type="button" class="btn-qty" onclick="stepProductQty(\'{prod["cod"]}\', \'cajas\', 1)">+</button>')
                html_out.append('                  </div>')
                html_out.append('                </div>')
                html_out.append('                <div class="qty-group" title="Escribe la cantidad de Unidades sueltas o usa + / −">')
                html_out.append('                  <span class="qty-label">Uni:</span>')
                html_out.append('                  <div class="product-qty-selector">')
                html_out.append(f'                    <button type="button" class="btn-qty" onclick="stepProductQty(\'{prod["cod"]}\', \'uni\', -1)">−</button>')
                html_out.append(f'                    <input type="number" id="uni_{prod["cod"]}" class="input-qty" value="0" min="0" oninput="onDirectInput(\'{prod["cod"]}\')" onchange="onDirectInput(\'{prod["cod"]}\')" data-code="{prod["cod"]}" data-name="{safe_name}" data-brand="{safe_brand}" data-unit="{safe_unit}" placeholder="0" />')
                html_out.append(f'                    <button type="button" class="btn-qty" onclick="stepProductQty(\'{prod["cod"]}\', \'uni\', 1)">+</button>')
                html_out.append('                  </div>')
                html_out.append('                </div>')
                html_out.append('              </div>')
                html_out.append('            </div>')
                
                html_out.append('          </div>') # card-body
                html_out.append('        </div>') # product-card
            html_out.append('      </div>') # products-grid
            html_out.append('    </div>') # category-section
            
        html_out.append('  </div>') # brand-section
        
    # ─── WIDGETS INTERACTIVOS DE PEDIDO (BARRA FLOTANTE Y MODAL) ───
    clean_biz_phone = re.sub(r'[^\d]', '', str(whatsapp_phone or ''))
    html_out.append('  <!-- Tarjeta de aviso de no resultados de búsqueda -->')
    html_out.append('  <div id="no-search-results-box" class="no-results-card" style="display: none;">')
    html_out.append('    <div style="font-size: 26pt; margin-bottom: 8px;">🔍</div>')
    html_out.append('    <div style="font-weight: 800; font-size: 11pt; color: #0F172A; margin-bottom: 4px;">No encontramos productos coincidentes</div>')
    html_out.append('    <div style="font-size: 8.5pt; color: #64748B; margin-bottom: 12px;">Intenta buscando con otra palabra, código o marca.</div>')
    html_out.append('    <button type="button" onclick="clearCatalogSearch()" style="background: #0F172A; color: #FFFFFF; border: none; padding: 7px 16px; border-radius: 20px; font-weight: 700; font-size: 8pt; cursor: pointer;">Ver todos los productos</button>')
    html_out.append('  </div>')
    
    html_out.append(f"""
  <!-- Barra Flotante de Pedido General -->
  <div id="floating-cart-bar" class="floating-cart-bar">
    <div class="cart-summary" onclick="openOrderModal()">
      <div class="cart-icon-wrapper">
        <svg viewBox="0 0 24 24" width="20" height="20" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M6 2L3 6v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2V6l-3-4z"></path><line x1="3" y1="6" x2="21" y2="6"></line><path d="M16 10a4 4 0 0 1-8 0"></path></svg>
        <span id="cart-badge" class="cart-badge">0</span>
      </div>
      <div class="cart-text">
        <span class="cart-title"><strong id="cart-qty-total">0 items</strong> seleccionados</span>
        <span class="cart-subtitle"><span id="cart-prod-total">0</span> producto(s) en tu lista</span>
      </div>
    </div>
    <div class="cart-actions">
      <button type="button" class="btn-view-order" onclick="openOrderModal()">
        <svg viewBox="0 0 24 24" width="14" height="14" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path><circle cx="12" cy="12" r="3"></circle></svg>
        <span>Ver Pedido</span>
      </button>
      <button type="button" class="btn-send-whatsapp-order" onclick="sendFullOrderWhatsApp()">
        <svg viewBox="0 0 24 24" width="15" height="15" fill="currentColor"><path d="M.057 24l1.687-6.163c-1.041-1.804-1.588-3.849-1.587-5.946.003-6.556 5.338-11.891 11.893-11.891 3.181.001 6.167 1.24 8.413 3.488 2.245 2.248 3.481 5.236 3.48 8.414-.003 6.557-5.338 11.892-11.893 11.892-1.99-.001-3.951-.5-5.688-1.448l-6.305 1.654zm6.597-3.807c1.676.995 3.276 1.591 5.392 1.592 5.448 0 9.886-4.434 9.889-9.885.002-5.462-4.415-9.89-9.881-9.892-5.452 0-9.887 4.434-9.889 9.884-.001 2.225.651 3.891 1.746 5.634l-.999 3.648 3.742-.981zm11.387-5.464c-.074-.124-.272-.198-.57-.347-.297-.149-1.758-.868-2.031-.967-.272-.099-.47-.149-.669.149-.198.297-.768.967-.941 1.165-.173.198-.347.223-.644.074-.297-.149-1.255-.462-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.297-.347.446-.521.151-.172.2-.296.3-.495.099-.198.05-.372-.025-.521-.075-.148-.669-1.611-.916-2.206-.242-.579-.487-.501-.669-.51l-.57-.01c-.198 0-.52.074-.792.372s-1.04 1.016-1.04 2.479 1.065 2.876 1.213 3.074c.149.198 2.095 3.2 5.076 4.487.709.306 1.263.489 1.694.626.712.226 1.36.194 1.872.118.571-.085 1.758-.719 2.006-1.413.248-.695.248-1.29.173-1.414z"/></svg>
        <span>Hacer Pedido</span>
      </button>
    </div>
  </div>

  <!-- Modal Resumen de Pedido -->
  <div id="order-modal-backdrop" class="modal-backdrop" onclick="closeOrderModal(event)">
    <div class="modal-card" onclick="event.stopPropagation()">
      <div class="modal-header">
        <div class="modal-title">
          <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M6 2L3 6v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2V6l-3-4z"></path><line x1="3" y1="6" x2="21" y2="6"></line><path d="M16 10a4 4 0 0 1-8 0"></path></svg>
          <span>Resumen de tu Pedido</span>
        </div>
        <button type="button" class="modal-close" onclick="closeOrderModal()">✕</button>
      </div>
      <div class="modal-body">
        <div class="customer-inputs">
          <input type="text" id="client-name" placeholder="Nombre / Empresa * (Requerido)" class="customer-field" required />
          <div class="customer-inputs-grid">
            <input type="text" id="client-address" placeholder="Dirección / Zona * (Requerido)" class="customer-field" required />
            <input type="text" id="client-phone" placeholder="Celular / WhatsApp (opcional)" class="customer-field" />
          </div>
        </div>
        <div id="order-items-list" class="order-items-list">
          <!-- Renderizado dinámico -->
        </div>
        <div class="order-total-banner">
          <span>Total en tu Pedido:</span>
          <strong id="modal-total-boxes">0 items</strong>
        </div>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn-clear-cart" onclick="clearCart()">Vaciar</button>
        <button type="button" class="btn-send-whatsapp-large" onclick="sendFullOrderWhatsApp()">
          <svg viewBox="0 0 24 24" width="15" height="15" fill="currentColor"><path d="M.057 24l1.687-6.163c-1.041-1.804-1.588-3.849-1.587-5.946.003-6.556 5.338-11.891 11.893-11.891 3.181.001 6.167 1.24 8.413 3.488 2.245 2.248 3.481 5.236 3.48 8.414-.003 6.557-5.338 11.892-11.893 11.892-1.99-.001-3.951-.5-5.688-1.448l-6.305 1.654zm6.597-3.807c1.676.995 3.276 1.591 5.392 1.592 5.448 0 9.886-4.434 9.889-9.885.002-5.462-4.415-9.89-9.881-9.892-5.452 0-9.887 4.434-9.889 9.884-.001 2.225.651 3.891 1.746 5.634l-.999 3.648 3.742-.981zm11.387-5.464c-.074-.124-.272-.198-.57-.347-.297-.149-1.758-.868-2.031-.967-.272-.099-.47-.149-.669.149-.198.297-.768.967-.941 1.165-.173.198-.347.223-.644.074-.297-.149-1.255-.462-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.297-.347.446-.521.151-.172.2-.296.3-.495.099-.198.05-.372-.025-.521-.075-.148-.669-1.611-.916-2.206-.242-.579-.487-.501-.669-.51l-.57-.01c-.198 0-.52.074-.792.372s-1.04 1.016-1.04 2.479 1.065 2.876 1.213 3.074c.149.198 2.095 3.2 5.076 4.487.709.306 1.263.489 1.694.626.712.226 1.36.194 1.872.118.571-.085 1.758-.719 2.006-1.413.248-.695.248-1.29.173-1.414z"/></svg>
          <span>Enviar Pedido por WhatsApp</span>
        </button>
      </div>
    </div>
  </div>

  <!-- Toast Flotante para Límites de Stock -->
  <div id="stock-toast-box" class="stock-toast"></div>

  <script>
    const BUSINESS_PHONE = "{clean_biz_phone}";
    const STOCK_API_URL = "{URL_STOCK_API}";
    let liveStockMap = {{}};
    const cart = {{}};
    let lastStockSyncTimestamp = 0;
    let isFetchingLiveStock = false;

    async function fetchLiveStock(isManual = false) {{
      if (isFetchingLiveStock && !isManual) return;
      isFetchingLiveStock = true;

      const statusEl = document.getElementById('live-stock-indicator');
      
      if (isManual && statusEl) {{
        statusEl.innerHTML = '<span class="pulse-dot-loading"></span> Sincronizando stock...';
      }}

      // 1. CARGA INMEDIATA DESDE CACHÉ LOCAL (0 milisegundos para render ultra-veloz)
      try {{
        const cached = localStorage.getItem('cached_stock_data');
        if (cached && (!liveStockMap || Object.keys(liveStockMap).length === 0)) {{
          const parsed = JSON.parse(cached);
          if (parsed && parsed.data) {{
            liveStockMap = parsed.data;
            applyStockData(liveStockMap);
            if (statusEl) {{
              const timeStr = parsed.timeStr || '';
              statusEl.innerHTML = `<span class="pulse-dot-online"></span> Stock en vivo${{timeStr ? ` (${{timeStr}})` : ''}}`;
            }}
          }}
        }}
      }} catch(e) {{}}

      if (statusEl && (!liveStockMap || Object.keys(liveStockMap).length === 0)) {{
        statusEl.innerHTML = '<span class="pulse-dot-loading"></span> Conectando almacén...';
      }}

      // 2. CONSULTA ULTRA RÁPIDA CON VERCEL EDGE CDN
      try {{
        let res = null;
        
        // Intento 1: Llamar a /api/stock (aprovecha Edge CDN en Vercel, o bypass si es manual)
        const apiUrl = isManual ? ('/api/stock?force=1&_t=' + Date.now()) : '/api/stock';
        try {{
          const c1 = new AbortController();
          const t1 = setTimeout(() => c1.abort(), 8000); // 8s máx en Edge
          res = await fetch(apiUrl, {{ signal: c1.signal }});
          clearTimeout(t1);
        }} catch (e) {{
          res = null;
        }}

        // Intento 2: Si /api/stock falló (ej: abriendo directo archivo file:///), consultar directo a Google Apps Script
        if (!res || !res.ok) {{
          try {{
            const c2 = new AbortController();
            const t2 = setTimeout(() => c2.abort(), 12000); // 12s máx directo
            const directUrl = STOCK_API_URL + (STOCK_API_URL.includes('?') ? '&' : '?') + '_t=' + Date.now();
            res = await fetch(directUrl, {{ 
              cache: 'no-store',
              redirect: 'follow',
              signal: c2.signal 
            }});
            clearTimeout(t2);
          }} catch (e) {{
            res = null;
          }}
        }}
        
        if (!res || !res.ok) throw new Error(res ? ('HTTP ' + res.status) : 'Sin respuesta de servidor');
        const data = await res.json();
        if (data && !data.error) {{
          liveStockMap = data;
          lastStockSyncTimestamp = Date.now();
          const now = new Date();
          const timeStr = now.toLocaleTimeString([], {{ hour: '2-digit', minute: '2-digit' }});
          
          try {{
            localStorage.setItem('cached_stock_data', JSON.stringify({{
              timestamp: Date.now(),
              timeStr: timeStr,
              data: liveStockMap
            }}));
          }} catch(e) {{}}
          
          applyStockData(liveStockMap);
          
          if (statusEl) {{
            statusEl.innerHTML = `<span class="pulse-dot-online"></span> Stock en vivo (${{timeStr}})`;
          }}
        }} else {{
          throw new Error(data.error || 'Respuesta inválida');
        }}
      }} catch (err) {{
        console.warn("Stock en vivo:", err.message);
        if (statusEl) {{
          const cached = localStorage.getItem('cached_stock_data');
          if (cached) {{
            try {{
              const parsed = JSON.parse(cached);
              if (parsed && parsed.timeStr) {{
                statusEl.innerHTML = `<span class="pulse-dot-online"></span> Stock en vivo (${{parsed.timeStr}})`;
                isFetchingLiveStock = false;
                return;
              }}
            }} catch(e) {{}}
          }}
          statusEl.innerHTML = '<span class="pulse-dot-online"></span> Stock en vivo';
        }}
      }} finally {{
        isFetchingLiveStock = false;
      }}
    }}

    function applyStockData(stockMap) {{
      if (!stockMap || typeof stockMap !== 'object') return;
      
      const cards = document.querySelectorAll('.product-card');
      cards.forEach(card => {{
        const rawCode = card.getAttribute('data-code') || '';
        const normCode = rawCode.toUpperCase().replace(/\\s+/g, '');
        const info = stockMap[normCode] || stockMap[rawCode.toUpperCase()];
        
        const pillEl = card.querySelector('.stock-status-pill') || document.getElementById(`stock_pill_${{rawCode}}`);
        const pkgEl = card.querySelector('.packaging-info') || document.getElementById(`pkg_info_${{rawCode}}`);

        if (info) {{
          const cantCaja = info.c || info.cantPorCaja || 1;
          const unMed = info.u || info.unidadMedida || "UNI";
          if (pkgEl) {{
            pkgEl.innerHTML = `📦 ${{cantCaja}} ${{unMed}} / Caja`;
            pkgEl.setAttribute('title', `Viene ${{cantCaja}} ${{unMed}} por caja`);
          }}
          
          const stock = typeof info.s === 'number' ? info.s : (typeof info.stockActual === 'number' ? info.stockActual : 0);
          const cajas = typeof info.b === 'number' ? info.b : (typeof info.cajas === 'number' ? info.cajas : Math.floor(stock / cantCaja));
          const estado = info.e || info.estado || "AGOTADO";
          
          if (pillEl) {{
            if (stock <= 0 || estado === "AGOTADO") {{
              pillEl.className = "stock-status-pill stock-out";
              pillEl.innerHTML = "🔴 Agotado";
              pillEl.setAttribute('title', "Sin stock disponible en almacén");
              card.classList.add('is-out-of-stock');
            }} else if (estado === "POCO_STOCK" || cajas <= 3) {{
              pillEl.className = "stock-status-pill stock-low";
              if (cajas >= 1) {{
                pillEl.innerHTML = `🟡 ¡Últimas ${{cajas}} caja${{cajas > 1 ? 's' : ''}}!`;
              }} else {{
                pillEl.innerHTML = `🟡 ¡Últimas ${{stock}} ${{unMed}}!`;
              }}
              pillEl.setAttribute('title', `Stock: ${{stock}} ${{unMed}} (${{cajas}} cajas) en almacén`);
              card.classList.remove('is-out-of-stock');
            }} else {{
              pillEl.className = "stock-status-pill stock-in-stock";
              pillEl.innerHTML = "🟢 En Stock";
              pillEl.setAttribute('title', "Stock disponible");
              card.classList.remove('is-out-of-stock');
            }}
          }}
        }} else {{
          if (pillEl) {{
            pillEl.className = "stock-status-pill stock-in-stock";
            pillEl.innerHTML = "🟢 Disponible";
          }}
        }}
      }});
    }}

    // Iniciar sincronización automática directa
    try {{
      if (document.readyState === 'loading') {{
        document.addEventListener('DOMContentLoaded', () => {{
          fetchLiveStock(false);
        }});
      }} else {{
        fetchLiveStock(false);
      }}
      // Intervalo regular silencioso cada 30 segundos
      setInterval(() => fetchLiveStock(false), 30000);

      // Auto-refresco al recuperar el foco de la pestaña si pasaron >30s
      document.addEventListener('visibilitychange', () => {{
        if (document.visibilityState === 'visible' && (Date.now() - lastStockSyncTimestamp > 30000)) {{
          fetchLiveStock(false);
        }}
      }});
      window.addEventListener('focus', () => {{
        if (Date.now() - lastStockSyncTimestamp > 30000) {{
          fetchLiveStock(false);
        }}
      }});
    }} catch(e) {{}}
    function quitarProductoEnVivo(e, code) {{
      if (e) {{
        e.stopPropagation();
        e.preventDefault();
      }}
      const upper = (code || '').toUpperCase();
      const card = document.querySelector(`.product-card[data-code="${{upper}}"]`) || (e ? e.target.closest('.product-card') : null);
      if (card) {{
        card.style.transition = 'all 0.3s cubic-bezier(0.4, 0, 0.2, 1)';
        card.style.opacity = '0';
        card.style.transform = 'scale(0.8)';
        setTimeout(() => {{
          card.style.display = 'none';
          filterCatalogLive(document.getElementById('catalog-live-search')?.value || '');
        }}, 300);
      }}
      try {{
        if (window.parent && window.parent !== window) {{
          window.parent.postMessage({{ type: 'REMOVE_CATALOG_ITEM', code: upper }}, '*');
        }}
      }} catch(err) {{}}
    }}

    function filterCatalogLive(query) {{
      const q = (query || '').trim().toUpperCase();
      const clearBtn = document.getElementById('catalog-search-clear-btn');
      if (clearBtn) clearBtn.style.display = q ? 'flex' : 'none';

      const words = q ? q.split(/\\s+/).filter(w => w.length > 0) : [];
      const cards = document.querySelectorAll('.product-card');
      let visibleCount = 0;

      const activeBrand = (currentBrandFilter || 'all').trim().toUpperCase();

      cards.forEach(card => {{
        const searchTarget = (card.getAttribute('data-search') || '').toUpperCase();
        const cardBrand = (card.getAttribute('data-brand') || '').trim().toUpperCase();

        const matchesBrand = (activeBrand === 'ALL' || cardBrand === activeBrand);
        const matchesQuery = (words.length === 0) || words.every(w => searchTarget.includes(w));

        const isVisible = matchesBrand && matchesQuery;
        card.style.display = isVisible ? '' : 'none';
        if (isVisible) visibleCount++;
      }});

      // Ocultar categorías sin productos visibles
      document.querySelectorAll('.category-section').forEach(sec => {{
        const hasVisible = Array.from(sec.querySelectorAll('.product-card')).some(c => c.style.display !== 'none');
        sec.style.display = hasVisible ? '' : 'none';
      }});

      // Ocultar marcas sin productos visibles
      document.querySelectorAll('.brand-section').forEach(sec => {{
        const hasVisible = Array.from(sec.querySelectorAll('.product-card')).some(c => c.style.display !== 'none');
        sec.style.display = hasVisible ? '' : 'none';
      }});

      const countEl = document.getElementById('search-match-count');
      if (countEl) countEl.innerText = visibleCount;

      const noResultsBox = document.getElementById('no-search-results-box');
      if (noResultsBox) noResultsBox.style.display = (visibleCount === 0) ? 'block' : 'none';
    }}

    function clearCatalogSearch() {{
      const input = document.getElementById('catalog-live-search');
      if (input) input.value = '';
      currentBrandFilter = 'all';
      document.querySelectorAll('.brand-chip').forEach(btn => {{
        const chipVal = btn.getAttribute('data-brand-chip') || '';
        btn.classList.toggle('active', chipVal.toUpperCase() === 'ALL');
      }});
      filterCatalogLive('');
    }}

    function filterByBrand(brand) {{
      currentBrandFilter = brand;
      const targetBrand = (brand || 'all').trim().toUpperCase();
      document.querySelectorAll('.brand-chip').forEach(btn => {{
        const chipVal = (btn.getAttribute('data-brand-chip') || '').trim().toUpperCase();
        btn.classList.toggle('active', chipVal === targetBrand);
      }});
      const input = document.getElementById('catalog-live-search');
      filterCatalogLive(input ? input.value : '');

      if (targetBrand !== 'ALL') {{
        const brandSecs = document.querySelectorAll('.brand-section');
        for (let i = 0; i < brandSecs.length; i++) {{
          const sec = brandSecs[i];
          const secBrand = (sec.getAttribute('data-brand-section') || '').trim().toUpperCase();
          if (secBrand === targetBrand) {{
            const stickyBar = document.getElementById('catalog-search-sticky-bar');
            const offset = stickyBar ? stickyBar.offsetHeight + 24 : 90;
            const pos = sec.getBoundingClientRect().top + window.pageYOffset - offset;
            window.scrollTo({{ top: pos, behavior: 'smooth' }});
            break;
          }}
        }}
      }}
    }}

    let toastTimeout = null;
    function showStockToast(msg) {{
      const box = document.getElementById('stock-toast-box');
      if (!box) return;
      box.innerHTML = msg;
      box.classList.add('show');
      if (toastTimeout) clearTimeout(toastTimeout);
      toastTimeout = setTimeout(() => {{
        box.classList.remove('show');
      }}, 3500);
    }}

    function highlightInputLimit(input) {{
      if (!input) return;
      input.classList.add('input-qty-error');
      setTimeout(() => input.classList.remove('input-qty-error'), 800);
    }}

    function stepProductQty(code, type, delta) {{
      const input = document.getElementById(type + '_' + code);
      if (!input) return;
      
      const normCode = (code || '').toUpperCase().replace(/\\s+/g, '');
      const stockInfo = liveStockMap[normCode] || liveStockMap[(code || '').toUpperCase()];
      
      let currentVal = parseInt(input.value) || 0;
      let desiredVal = Math.max(0, currentVal + delta);

      if (stockInfo) {{
        const cantCaja = stockInfo.c || stockInfo.cantPorCaja || 1;
        const stockActual = typeof stockInfo.s === 'number' ? stockInfo.s : (typeof stockInfo.stockActual === 'number' ? stockInfo.stockActual : 0);
        const maxCajas = typeof stockInfo.b === 'number' ? stockInfo.b : (typeof stockInfo.cajas === 'number' ? stockInfo.cajas : Math.floor(stockActual / cantCaja));
        const unMed = stockInfo.u || stockInfo.unidadMedida || 'UNI';
        const estado = stockInfo.e || stockInfo.estado || 'AGOTADO';

        if (stockActual <= 0 || estado === 'AGOTADO') {{
          input.value = 0;
          showStockToast(`🔴 El producto <strong>${{code}}</strong> está agotado.`);
          onDirectInput(code);
          return;
        }}

        if (type === 'cajas' && delta > 0) {{
          const currentUni = parseInt(document.getElementById('uni_' + code)?.value) || 0;
          const totalUnitsIfAdded = desiredVal * cantCaja + currentUni;
          
          if (desiredVal > maxCajas || totalUnitsIfAdded > stockActual) {{
            input.value = maxCajas;
            showStockToast(`⚠️ Stock límite: solo quedan <strong>${{maxCajas}} caja${{maxCajas > 1 ? 's' : ''}}</strong> de ${{code}}.`);
            highlightInputLimit(input);
            onDirectInput(code);
            return;
          }}
        }} else if (type === 'uni' && delta > 0) {{
          const currentCajas = parseInt(document.getElementById('cajas_' + code)?.value) || 0;
          const totalUnitsIfAdded = currentCajas * cantCaja + desiredVal;
          
          if (totalUnitsIfAdded > stockActual) {{
            const remainingUni = Math.max(0, stockActual - currentCajas * cantCaja);
            input.value = remainingUni;
            showStockToast(`⚠️ Stock límite: solo quedan <strong>${{remainingUni}} ${{unMed}}</strong> adicionales.`);
            highlightInputLimit(input);
            onDirectInput(code);
            return;
          }}
        }}
      }}

      input.value = desiredVal;
      onDirectInput(code);
    }}

    function onDirectInput(code) {{
      const inputCajas = document.getElementById('cajas_' + code);
      const inputUni = document.getElementById('uni_' + code);
      const ref = inputCajas || inputUni;
      if (!ref) return;

      let cajas = parseInt(inputCajas?.value) || 0;
      let uni = parseInt(inputUni?.value) || 0;
      cajas = Math.max(0, cajas);
      uni = Math.max(0, uni);

      const normCode = (code || '').toUpperCase().replace(/\\s+/g, '');
      const stockInfo = liveStockMap[normCode] || liveStockMap[(code || '').toUpperCase()];
      const cantCaja = (stockInfo && (stockInfo.c || stockInfo.cantPorCaja)) ? (stockInfo.c || stockInfo.cantPorCaja) : (parseInt(ref.getAttribute('data-caja')) || 1);
      const unMed = (stockInfo && (stockInfo.u || stockInfo.unidadMedida)) ? (stockInfo.u || stockInfo.unidadMedida) : (ref.getAttribute('data-unit') || 'UNI');
      const maxCajas = stockInfo ? (typeof stockInfo.b === 'number' ? stockInfo.b : (typeof stockInfo.cajas === 'number' ? stockInfo.cajas : Math.floor((stockInfo.s || 0) / cantCaja))) : 99999;
      const stockActual = stockInfo ? (typeof stockInfo.s === 'number' ? stockInfo.s : (typeof stockInfo.stockActual === 'number' ? stockInfo.stockActual : 99999)) : 99999;
      const estado = stockInfo ? (stockInfo.e || stockInfo.estado || 'DISPONIBLE') : 'DISPONIBLE';

      if (stockInfo && (stockActual <= 0 || estado === 'AGOTADO')) {{
        if (cajas > 0 || uni > 0) {{
          showStockToast(`🔴 El producto <strong>${{code}}</strong> está agotado.`);
        }}
        cajas = 0;
        uni = 0;
        if (inputCajas) inputCajas.value = 0;
        if (inputUni) inputUni.value = 0;
      }} else {{
        // 1. AUTO-CONVERSIÓN INTELIGENTE EN TIEMPO REAL: Si las unidades completan una o más cajas
        if (cantCaja > 1 && uni >= cantCaja) {{
          const extraCajas = Math.floor(uni / cantCaja);
          const remainingUni = uni % cantCaja;
          const targetCajas = cajas + extraCajas;
          
          if (targetCajas <= maxCajas) {{
            cajas = targetCajas;
            uni = remainingUni;
            if (inputCajas) inputCajas.value = cajas;
            if (inputUni) inputUni.value = uni;
            showStockToast(`📦 <strong>${{extraCajas * cantCaja}} unidades</strong> convertidas a <strong>${{extraCajas}} caja${{extraCajas > 1 ? 's' : ''}}</strong> cerrada${{extraCajas > 1 ? 's' : ''}}.`);
            highlightInputLimit(inputCajas);
          }} else {{
            const possibleCajas = Math.max(0, maxCajas - cajas);
            if (possibleCajas > 0) {{
              cajas = maxCajas;
              uni = Math.max(0, uni - possibleCajas * cantCaja);
              if (inputCajas) inputCajas.value = cajas;
              if (inputUni) inputUni.value = uni;
              showStockToast(`📦 Convertidas <strong>${{possibleCajas}} cajas</strong> (límite de stock).`);
              highlightInputLimit(inputCajas);
            }}
          }}
        }}

        // 2. SUGERENCIA INTELIGENTE: Si le falta solo 1 unidad para completar caja
        if (cantCaja > 1 && uni > 0 && (cantCaja - uni === 1) && (cajas + 1 <= maxCajas)) {{
          showStockToast(`💡 ¡Agrega <strong>1 ${{unMed}}</strong> más para completar <strong>${{cajas + 1}} cajas</strong> cerradas!`);
        }}

        if (stockInfo) {{
          if (cajas > maxCajas) {{
            cajas = maxCajas;
            if (inputCajas) inputCajas.value = maxCajas;
            showStockToast(`⚠️ Stock ajustado: máximo <strong>${{maxCajas}} caja${{maxCajas > 1 ? 's' : ''}}</strong>.`);
            highlightInputLimit(inputCajas);
          }}
          let totalUniSelected = cajas * cantCaja + uni;
          if (totalUniSelected > stockActual) {{
            uni = Math.max(0, stockActual - cajas * cantCaja);
            if (inputUni) inputUni.value = uni;
            showStockToast(`⚠️ Stock ajustado: solo quedan <strong>${{uni}} ${{unMed}}</strong> sueltas.`);
            highlightInputLimit(inputUni);
          }}
        }}
      }}

      if (inputCajas && inputCajas.value !== '' && parseInt(inputCajas.value) < 0) inputCajas.value = 0;
      if (inputUni && inputUni.value !== '' && parseInt(inputUni.value) < 0) inputUni.value = 0;

      const totalUnitsThisItem = cajas * cantCaja + uni;

      const card = ref.closest('.product-card');
      if (cajas > 0 || uni > 0) {{
        cart[code] = {{
          code: code,
          cajas: cajas,
          uni: uni,
          cantPorCaja: cantCaja,
          totalUnits: totalUnitsThisItem,
          name: ref.getAttribute('data-name') || code,
          brand: ref.getAttribute('data-brand') || '',
          unit: unMed
        }};
        if (card) card.classList.add('has-ordered');
      }} else {{
        delete cart[code];
        if (card) card.classList.remove('has-ordered');
      }}

      updateCartUI();
    }}

    function updateCartUI() {{
      const items = Object.values(cart);
      const totalProds = items.length;
      let totalCajas = 0;
      let totalUni = 0;
      let grandTotalPieces = 0;
      items.forEach(it => {{
        const normCode = (it.code || '').toUpperCase().replace(/\\s+/g, '');
        const stockInfo = liveStockMap[normCode] || liveStockMap[(it.code || '').toUpperCase()];
        const cantCaja = stockInfo ? (stockInfo.c || stockInfo.cantPorCaja || it.cantPorCaja || 1) : (it.cantPorCaja || 1);
        totalCajas += it.cajas;
        totalUni += it.uni;
        grandTotalPieces += (it.cajas * cantCaja + it.uni);
      }});

      const cartBar = document.getElementById('floating-cart-bar');
      const badge = document.getElementById('cart-badge');
      const qtyTotalEl = document.getElementById('cart-qty-total');
      const prodTotalEl = document.getElementById('cart-prod-total');
      const modalTotalBoxes = document.getElementById('modal-total-boxes');

      if (badge) badge.innerText = totalProds;
      
      let summaryText = [];
      if (totalCajas > 0) summaryText.push(totalCajas + ' caja' + (totalCajas > 1 ? 's' : ''));
      if (totalUni > 0) summaryText.push(totalUni + ' unid');
      
      const displayFloating = summaryText.length > 0 ? `${{summaryText.join(' + ')}} (= ${{grandTotalPieces}} unid)` : '0 items';
      const displayModal = summaryText.length > 0 ? `${{summaryText.join(' + ')}} (= ${{grandTotalPieces}} unidades en total)` : '0 items';

      if (qtyTotalEl) qtyTotalEl.innerText = displayFloating;
      if (prodTotalEl) prodTotalEl.innerText = totalProds;
      if (modalTotalBoxes) modalTotalBoxes.innerText = displayModal;

      if (totalCajas > 0 || totalUni > 0) {{
        if (cartBar) cartBar.classList.add('visible');
      }} else {{
        if (cartBar) cartBar.classList.remove('visible');
      }}

      renderModalList();
    }}

    function openOrderModal() {{
      const items = Object.values(cart);
      if (items.length === 0) {{
        alert("Aún no has seleccionado cantidades para ningún producto. Escribe o usa + / − en Cajas o Unidades.");
        return;
      }}
      
      // Cargar memoria del cliente si existe
      try {{
        const savedName = localStorage.getItem('last_client_name');
        const savedAddress = localStorage.getItem('last_client_address');
        const savedPhone = localStorage.getItem('last_client_phone');
        const nameEl = document.getElementById('client-name');
        const addrEl = document.getElementById('client-address');
        const phoneEl = document.getElementById('client-phone');
        if (savedName && nameEl && !nameEl.value) nameEl.value = savedName;
        if (savedAddress && addrEl && !addrEl.value) addrEl.value = savedAddress;
        if (savedPhone && phoneEl && !phoneEl.value) phoneEl.value = savedPhone;
      }} catch(e) {{}}

      renderModalList();
      const modal = document.getElementById('order-modal-backdrop');
      if (modal) modal.classList.add('open');
    }}

    function closeOrderModal(e) {{
      const modal = document.getElementById('order-modal-backdrop');
      if (modal) modal.classList.remove('open');
    }}

    function renderModalList() {{
      const container = document.getElementById('order-items-list');
      if (!container) return;
      const items = Object.values(cart);
      if (items.length === 0) {{
        container.innerHTML = '<div style="text-align: center; padding: 20px; color: #94A3B8; font-size: 8.5pt;">No tienes productos en tu pedido.</div>';
        return;
      }}

      let html = '';
      items.forEach(it => {{
        const normCode = (it.code || '').toUpperCase().replace(/\\s+/g, '');
        const stockInfo = liveStockMap[normCode] || liveStockMap[(it.code || '').toUpperCase()];
        const cantCaja = stockInfo ? (stockInfo.c || stockInfo.cantPorCaja || it.cantPorCaja || 1) : (it.cantPorCaja || 1);
        const unMed = stockInfo ? (stockInfo.u || stockInfo.unidadMedida || it.unit || 'UNI') : (it.unit || 'UNI');
        const itemTotalUnits = (it.cajas || 0) * cantCaja + (it.uni || 0);

        let breakdownText = '';
        if (it.cajas > 0 && it.uni > 0) {{
          breakdownText = `${{it.cajas}} caja${{it.cajas > 1 ? 's' : ''}} + ${{it.uni}} ${{unMed}}`;
        }} else if (it.cajas > 0) {{
          breakdownText = `${{it.cajas}} caja${{it.cajas > 1 ? 's' : ''}}`;
        }} else if (it.uni > 0) {{
          breakdownText = `${{it.uni}} ${{unMed}}`;
        }}

        html += `
          <div class="order-item-row">
            <div class="order-item-info">
              <div style="display: flex; align-items: center; gap: 8px; flex-wrap: wrap;">
                <span class="order-item-code">${{it.code}}</span>
                <span style="font-size: 8pt; font-weight: 700; color: #94A3B8; background: rgba(255, 255, 255, 0.08); padding: 2px 7px; border-radius: 6px; text-transform: uppercase;">${{it.brand}}</span>
                <span style="font-size: 7.5pt; color: #94A3B8; background: rgba(255, 255, 255, 0.04); padding: 2px 6px; border-radius: 4px; border: 1px dashed rgba(255, 255, 255, 0.14);">📦 ${{cantCaja}} ${{unMed}}/caja</span>
              </div>
              <div class="order-item-name" title="${{it.name}}">${{it.name}}</div>
              <div style="display: flex; align-items: center; gap: 6px; margin-top: 3px;">
                <span style="font-size: 8.5pt; font-weight: 800; color: #86EFAC; background: rgba(34, 197, 94, 0.14); border: 1px solid rgba(34, 197, 94, 0.3); padding: 2px 9px; border-radius: 6px;">
                  Total: ${{itemTotalUnits}} ${{unMed}} (${{breakdownText}})
                </span>
              </div>
            </div>
            <div class="order-selectors-dual">
              <div class="qty-group">
                <span class="qty-label">Caja</span>
                <div class="product-qty-selector">
                  <button type="button" class="btn-qty" onclick="stepProductQty('${{it.code}}', 'cajas', -1)">−</button>
                  <input type="number" class="input-qty" value="${{it.cajas}}" min="0" oninput="onModalDirectInput('${{it.code}}', 'cajas', this.value)" />
                  <button type="button" class="btn-qty" onclick="stepProductQty('${{it.code}}', 'cajas', 1)">+</button>
                </div>
              </div>
              <div class="qty-group">
                <span class="qty-label">Uni</span>
                <div class="product-qty-selector">
                  <button type="button" class="btn-qty" onclick="stepProductQty('${{it.code}}', 'uni', -1)">−</button>
                  <input type="number" class="input-qty" value="${{it.uni}}" min="0" oninput="onModalDirectInput('${{it.code}}', 'uni', this.value)" />
                  <button type="button" class="btn-qty" onclick="stepProductQty('${{it.code}}', 'uni', 1)">+</button>
                </div>
              </div>
            </div>
          </div>
        `;
      }});
      container.innerHTML = html;
    }}

    function onModalDirectInput(code, type, val) {{
      const cardInput = document.getElementById(type + '_' + code);
      if (cardInput) {{
        cardInput.value = Math.max(0, parseInt(val) || 0);
      }}
      onDirectInput(code);
    }}

    function clearCart() {{
      Object.keys(cart).forEach(code => {{
        const inputCajas = document.getElementById('cajas_' + code);
        const inputUni = document.getElementById('uni_' + code);
        if (inputCajas) inputCajas.value = 0;
        if (inputUni) inputUni.value = 0;
        const ref = inputCajas || inputUni;
        const card = ref?.closest('.product-card');
        if (card) card.classList.remove('has-ordered');
      }});
      for (const k in cart) delete cart[k];
      updateCartUI();
      closeOrderModal();
    }}

    function sendFullOrderWhatsApp() {{
      const items = Object.values(cart);
      if (items.length === 0) {{
        alert("Por favor selecciona al menos 1 producto con su cantidad para enviar el pedido.");
        return;
      }}

      const nameInput = document.getElementById('client-name');
      const addrInput = document.getElementById('client-address');
      const clientName = (nameInput?.value || '').trim();
      const clientAddress = (addrInput?.value || '').trim();
      const clientPhone = (document.getElementById('client-phone')?.value || '').trim();

      if (!clientName) {{
        if (nameInput) {{
          nameInput.focus();
          nameInput.classList.add('input-qty-error');
          setTimeout(() => nameInput.classList.remove('input-qty-error'), 1200);
        }}
        showStockToast("⚠️ Por favor ingresa tu <strong>Nombre o Empresa</strong> para procesar tu pedido.");
        return;
      }}

      if (!clientAddress) {{
        if (addrInput) {{
          addrInput.focus();
          addrInput.classList.add('input-qty-error');
          setTimeout(() => addrInput.classList.remove('input-qty-error'), 1200);
        }}
        showStockToast("⚠️ Por favor ingresa tu <strong>Dirección o Zona</strong> para el despacho.");
        return;
      }}

      // Guardar memoria del cliente
      try {{
        if (clientName) localStorage.setItem('last_client_name', clientName);
        if (clientAddress) localStorage.setItem('last_client_address', clientAddress);
        if (clientPhone) localStorage.setItem('last_client_phone', clientPhone);
      }} catch(e) {{}}

      let totalCajas = 0;
      let totalUniLoose = 0;
      let grandTotalPieces = 0;
      items.forEach(it => {{
        const normCode = (it.code || '').toUpperCase().replace(/\\s+/g, '');
        const stockInfo = liveStockMap[normCode] || liveStockMap[(it.code || '').toUpperCase()];
        const cantCaja = stockInfo ? (stockInfo.c || stockInfo.cantPorCaja || it.cantPorCaja || 1) : (it.cantPorCaja || 1);
        totalCajas += it.cajas;
        totalUniLoose += it.uni;
        grandTotalPieces += (it.cajas * cantCaja + it.uni);
      }});

      let msg = "*📋 SOLICITUD DE PEDIDO - IMPORTADORA RIVERO*\\n";
      msg += "----------------------------------------\\n";
      if (clientName) msg += `👤 *Cliente:* ${{clientName}}\\n`;
      if (clientAddress) msg += `📍 *Dirección:* ${{clientAddress}}\\n`;
      if (clientPhone) msg += `📱 *Teléfono:* ${{clientPhone}}\\n`;
      msg += `📅 *Fecha:* ${{new Date().toLocaleDateString('es-ES')}}\\n`;
      msg += "----------------------------------------\\n\\n";

      msg += "*📦 DETALLE DEL PEDIDO:*\\n";
      items.forEach((it, idx) => {{
        const normCode = (it.code || '').toUpperCase().replace(/\\s+/g, '');
        const stockInfo = liveStockMap[normCode] || liveStockMap[(it.code || '').toUpperCase()];
        const cantCaja = stockInfo ? (stockInfo.c || stockInfo.cantPorCaja || it.cantPorCaja || 1) : (it.cantPorCaja || 1);
        const unMed = stockInfo ? (stockInfo.u || stockInfo.unidadMedida || it.unit || 'UNI') : (it.unit || 'UNI');
        const itemTotalUnits = (it.cajas || 0) * cantCaja + (it.uni || 0);

        let cantParts = [];
        if (it.cajas > 0) cantParts.push(`*${{it.cajas}} Cajas*`);
        if (it.uni > 0) cantParts.push(`*${{it.uni}} ${{unMed}}*`);
        const cantDesglose = cantParts.join(' + ');

        msg += `${{idx + 1}}. [${{it.code}}] ${{it.name}}\\n`;
        msg += `   ▪ Pedido: ${{cantDesglose}}  ➜  *Total: ${{itemTotalUnits}} ${{unMed}}*\\n`;
        msg += `   ▪ Empaque: ${{cantCaja}} ${{unMed}}/caja`;
        if (it.brand) msg += ` | Marca: ${{it.brand}}`;
        msg += "\\n\\n";
      }});

      msg += "----------------------------------------\\n";
      msg += "*📊 RESUMEN GENERAL:*\\n";
      if (totalCajas > 0) msg += `• Total Cajas cerradas: *${{totalCajas}}*\\n`;
      if (totalUniLoose > 0) msg += `• Total Unidades sueltas: *${{totalUniLoose}}*\\n`;
      msg += `• 📦 *TOTAL MERCADERÍA: ${{grandTotalPieces}} piezas / unidades*\\n`;
      msg += "----------------------------------------\\n";
      msg += "_Por favor confirmar disponibilidad y cotización. ¡Muchas gracias!_";

      const encoded = encodeURIComponent(msg);
      let targetUrl = '';
      if (BUSINESS_PHONE) {{
        targetUrl = `https://wa.me/${{BUSINESS_PHONE}}?text=${{encoded}}`;
      }} else {{
        targetUrl = `https://api.whatsapp.com/send?text=${{encoded}}`;
      }}

      window.open(targetUrl, '_blank');
    }}
  </script>
</body>
</html>""")
    
    html_path = output_filename
    with open(html_path, "w", encoding="utf-8") as f_html:
        f_html.write("\n".join(html_out))
        
    if hashes_updated:
        try:
            with open(hashes_path, "w", encoding="utf-8") as f_h:
                json.dump(image_hashes, f_h, indent=2)
            print(f"  [CACHE] Guardados nuevos hashes de imágenes en '{hashes_path}'")
        except Exception as e:
            print(f"  [AVISO] No se pudo guardar el archivo de hashes de imágenes: {e}")
        
    if no_encontrados:
        print(f"\n  [AVISO] {len(no_encontrados)} producto(s) NO encontrado(s) en la base de datos:")
        for no_cod in no_encontrados:
            print(f"    - Código: {no_cod}")
    else:
        print(f"\n  [OK] ¡Los {len(codigos)} productos solicitados fueron encontrados y generados con éxito!")
        
    return html_path, total_prods

def generar(descargar_nube=True, codigos_custom=None, layout="desktop", forzar_imagenes=False, whatsapp_phone=None):
    # Copiar Excel local de la raíz si existe y es más nuevo que la caché
    local_root_excel = "catalogos.xlsx"
    if os.path.exists(local_root_excel):
        if not os.path.exists(ARCHIVO_EXCEL) or os.path.getmtime(local_root_excel) > os.path.getmtime(ARCHIVO_EXCEL):
            print(f"\n[LOCAL] Detectado '{local_root_excel}' en la raíz más reciente que la caché. Copiando...")
            import shutil
            try:
                dest_dir = os.path.dirname(ARCHIVO_EXCEL)
                if dest_dir and not os.path.exists(dest_dir):
                    os.makedirs(dest_dir)
                shutil.copy2(local_root_excel, ARCHIVO_EXCEL)
                print(f"[LOCAL] [OK] Excel de la raíz copiado con éxito a la caché.")
            except Exception as e:
                print(f"[LOCAL] [AVISO] No se pudo copiar '{local_root_excel}': {e}")

    if descargar_nube and URL_GOOGLE_SHEETS:
        descargar_base_de_datos_nube(URL_GOOGLE_SHEETS, ARCHIVO_EXCEL)

    if not os.path.exists(ARCHIVO_EXCEL):
        print(f"\nERROR: No se encontró '{ARCHIVO_EXCEL}'")
        raise FileNotFoundError(f"No se encontró el archivo base de datos Excel: {ARCHIVO_EXCEL}")

    print(f"\nAbriendo {ARCHIVO_EXCEL}...")
    wb = load_workbook(ARCHIVO_EXCEL, data_only=True)

    # 1. Leer códigos a procesar
    codigos = []
    if codigos_custom:
        if isinstance(codigos_custom, str):
            tokens = re.split(r'[\r\n,;\t]+', codigos_custom)
            codigos = [normalizar_codigo(c) for c in tokens if normalizar_codigo(c)]
        elif isinstance(codigos_custom, (list, tuple)):
            for item in codigos_custom:
                if isinstance(item, str):
                    tokens = re.split(r'[\r\n,;\t]+', item)
                    for t in tokens:
                        c_clean = normalizar_codigo(t)
                        if c_clean:
                            codigos.append(c_clean)
                else:
                    c_clean = normalizar_codigo(item)
                    if c_clean:
                        codigos.append(c_clean)

    if not codigos:
        # Buscar en hoja Vista_Catalogo si existe
        ws_vista = None
        for name in wb.sheetnames:
            if "VISTA" in name.upper():
                ws_vista = wb[name]
                break
        if not ws_vista and HOJA_VISTA in wb.sheetnames:
            ws_vista = wb[HOJA_VISTA]
            
        if ws_vista:
            empty_count = 0
            max_vista_rows = min(ws_vista.max_row + 10, 5000) if ws_vista.max_row else 1000
            for row in range(FILA_INICIO_CODIGOS, max_vista_rows):
                val = ws_vista.cell(row=row, column=COLUMNA_CODIGOS).value
                if val and str(val).strip():
                    codigos.append(normalizar_codigo(val))
                    empty_count = 0
                else:
                    empty_count += 1
                    if empty_count >= 50:
                        break

    if not codigos:
        # Si no se ingresaron códigos manuales ni en Vista_Catalogo, procesar automáticamente todo el inventario
        print(">>> No se ingresaron códigos específicos: Cargando todo el inventario...")
        hojas_temp = detectar_hojas_inventario(wb)
        vistos_set = set()
        for ws_cur in hojas_temp:
            cols_cfg, fila_inicio = detectar_columnas(ws_cur)
            col_c = cols_cfg["codigo"]
            consecutive_empty = 0
            max_r = min(ws_cur.max_row + 50, 30000) if ws_cur.max_row else 15000
            for row in range(fila_inicio, max_r):
                val = ws_cur.cell(row=row, column=col_c).value
                if val and str(val).strip():
                    c_norm = normalizar_codigo(val)
                    if c_norm and c_norm.upper() not in vistos_set:
                        vistos_set.add(c_norm.upper())
                        codigos.append(c_norm)
                    consecutive_empty = 0
                else:
                    consecutive_empty += 1
                    if consecutive_empty >= 100:
                        break

    if not codigos:
        print(f"No se encontraron códigos de producto a procesar.")
        raise ValueError(f"No se encontraron códigos de producto en las hojas de inventario ni en Vista_Catalogo.")

    print(f"Códigos a procesar: {len(codigos)}")

    # 2. Identificar hojas de inventario y leer base de datos de productos de forma rápida y completa
    db = {}
    db_norm = {}
    db_clean = {}
    hojas_inv = detectar_hojas_inventario(wb)

    print(f"Hojas de inventario identificadas: {[ws.title for ws in hojas_inv]}")

    for ws_cur in hojas_inv:
        cols_cfg, fila_inicio = detectar_columnas(ws_cur)
        col_c = cols_cfg["codigo"]
        col_cat = cols_cfg["categoria"]
        col_tip = cols_cfg["tipo"]
        col_nom = cols_cfg["nombre"]
        col_siz = cols_cfg["size"]
        col_det = cols_cfg["detalle"]
        col_uni = cols_cfg["uni"]
        
        consecutive_empty = 0
        max_r = min(ws_cur.max_row + 50, 30000) if ws_cur.max_row else 15000
        
        for row in range(fila_inicio, max_r):
            cod_val = ws_cur.cell(row=row, column=col_c).value
            if cod_val is None or str(cod_val).strip() == "":
                consecutive_empty += 1
                if consecutive_empty >= 150:
                    break
                continue
            
            consecutive_empty = 0
            raw_clave = normalizar_codigo(cod_val)
            if not raw_clave:
                continue
                
            norm_k = raw_clave.upper()
            clean_k = clave_busqueda(raw_clave)
            
            col_cj = cols_cfg.get("cant_caja")
            cant_caja_val = 1
            if col_cj:
                try:
                    raw_cj = ws_cur.cell(row=row, column=col_cj).value
                    if raw_cj is not None:
                        cant_caja_val = float(str(raw_cj).replace(",", "."))
                        if cant_caja_val <= 0:
                            cant_caja_val = 1
                except Exception:
                    cant_caja_val = 1
            
            prod_info = {
                "categoria": str(ws_cur.cell(row=row, column=col_cat).value or "").strip(),
                "cod":       raw_clave,
                "tipo":      str(ws_cur.cell(row=row, column=col_tip).value or "").strip(),
                "nombre":    str(ws_cur.cell(row=row, column=col_nom).value or "").strip(),
                "size":      str(ws_cur.cell(row=row, column=col_siz).value or "").strip(),
                "detalle":   str(ws_cur.cell(row=row, column=col_det).value or "").strip(),
                "uni":       str(ws_cur.cell(row=row, column=col_uni).value or "pcs").strip(),
                "cant_caja": cant_caja_val,
                "fila_db":   row,
                "ws_title":  ws_cur.title,
            }
            
            if raw_clave not in db:
                db[raw_clave] = prod_info
            if norm_k not in db_norm:
                db_norm[norm_k] = prod_info
            if clean_k and clean_k not in db_clean:
                db_clean[clean_k] = prod_info

    print(f"Productos cargados en base de datos: {len(db)}")

    # 3. Determinar las filas de interés por hoja y extraer imágenes de forma optimizada
    filas_por_hoja = {ws_cur.title: set() for ws_cur in hojas_inv}
    
    for cod in codigos:
        prod = buscar_producto_en_db(cod, db, db_norm, db_clean)
        if prod:
            ws_name = prod.get("ws_title", hojas_inv[0].title)
            r = prod["fila_db"]
            if ws_name in filas_por_hoja:
                filas_por_hoja[ws_name].add(r)
                filas_por_hoja[ws_name].add(r - 1)
                filas_por_hoja[ws_name].add(r + 1)

    imagenes_por_fila = {}
    for ws_cur in hojas_inv:
        filas_int = filas_por_hoja.get(ws_cur.title, set())
        if filas_int:
            imgs = extraer_imagenes_db(ws_cur, filas_interes=filas_int)
            for r, img_data in imgs.items():
                imagenes_por_fila[(ws_cur.title, r)] = img_data
                if r not in imagenes_por_fila:
                    imagenes_por_fila[r] = img_data

    # 4. Eliminar hoja CATALOGO del Excel para reducir drásticamente el tamaño del archivo
    if HOJA_CATALOGO in wb.sheetnames:
        print(f"Limpiando hoja de catálogo antigua del Excel para reducir espacio...")
        del wb[HOJA_CATALOGO]
        try:
            wb.save(ARCHIVO_EXCEL)
            print(f"Excel guardado y optimizado ({ARCHIVO_EXCEL})")
        except Exception as e:
            print(f"  [AVISO] No se pudo optimizar el tamaño de '{ARCHIVO_EXCEL}': {e}")

    # 5. Generar HTML y extraer fotos a disco (primero el layout seleccionado)
    html_file, total_prods = generar_html_y_imagenes(db, codigos, imagenes_por_fila, layout=layout, output_filename="catalogos.html", forzar_imagenes=forzar_imagenes, db_norm=db_norm, db_clean=db_clean, whatsapp_phone=whatsapp_phone)

    # También generar copias de ambos layouts
    try:
        otro_layout = "mobile" if layout == "desktop" else "desktop"
        otro_filename = f"catalogos_{otro_layout}.html"
        generar_html_y_imagenes(db, codigos, imagenes_por_fila, layout=otro_layout, output_filename=otro_filename, forzar_imagenes=forzar_imagenes, db_norm=db_norm, db_clean=db_clean, whatsapp_phone=whatsapp_phone)
        # Guardar el diseño actual con su nombre específico también
        generar_html_y_imagenes(db, codigos, imagenes_por_fila, layout=layout, output_filename=f"catalogos_{layout}.html", forzar_imagenes=forzar_imagenes, db_norm=db_norm, db_clean=db_clean, whatsapp_phone=whatsapp_phone)
        print(f"  [HTML] Generadas copias específicas: 'catalogos_desktop.html' y 'catalogos_mobile.html'")
    except Exception as e:
        print(f"  [AVISO] No se pudo generar la copia del diseño alternativo: {e}")

    # 6. Proceso completado exitosamente (Solo HTML)
    total_no_encontrados = len([c for c in codigos if not buscar_producto_en_db(c, db, db_norm, db_clean)])
    print(f"\n[OK] ¡Proceso completado exitosamente!")
    print(f"  Catálogo HTML: {html_file}")
    print(f"  Total productos generados: {total_prods}")
    if total_no_encontrados > 0:
        print(f"  Total productos NO encontrados: {total_no_encontrados} (revisa el listado de avisos arriba)")
    else:
        print(f"  Todos los productos fueron encontrados correctamente.")

def obtener_resumen_inventario():
    """
    Lee el archivo Excel disponible y retorna un resumen ligero de productos, marcas y categorías
    para alimentar el buscador predictivo y los filtros del panel web.
    """
    excel_path = ARCHIVO_EXCEL
    if not os.path.exists(excel_path):
        if os.path.exists("catalogos.xlsx"):
            excel_path = "catalogos.xlsx"
        else:
            return {"productos": [], "marcas": {}, "categorias": {}, "total": 0}
            
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception:
        return {"productos": [], "marcas": {}, "categorias": {}, "total": 0}
        
    hojas_inv = detectar_hojas_inventario(wb)
    productos = []
    marcas = {}
    categorias = {}
    vistos = set()
    
    for ws_cur in hojas_inv:
        cols_cfg, fila_inicio = detectar_columnas(ws_cur)
        col_c = cols_cfg["codigo"]
        col_cat = cols_cfg["categoria"]
        col_tip = cols_cfg["tipo"]
        col_nom = cols_cfg["nombre"]
        col_siz = cols_cfg["size"]
        col_det = cols_cfg["detalle"]
        col_uni = cols_cfg["uni"]
        
        consecutive_empty = 0
        max_r = min(ws_cur.max_row + 50, 30000) if ws_cur.max_row else 15000
        
        for row in range(fila_inicio, max_r):
            cod_val = ws_cur.cell(row=row, column=col_c).value
            if cod_val is None or str(cod_val).strip() == "":
                consecutive_empty += 1
                if consecutive_empty >= 150:
                    break
                continue
            
            consecutive_empty = 0
            cod_str = normalizar_codigo(cod_val)
            if not cod_str or cod_str.upper() in vistos:
                continue
            vistos.add(cod_str.upper())
            
            nombre_str = str(ws_cur.cell(row=row, column=col_nom).value or "").strip()
            cat_str = str(ws_cur.cell(row=row, column=col_cat).value or "Sin Categoría").strip() or "Sin Categoría"
            tipo_str = str(ws_cur.cell(row=row, column=col_tip).value or "OTRO").strip() or "OTRO"
            size_str = str(ws_cur.cell(row=row, column=col_siz).value or "").strip()
            uni_str = str(ws_cur.cell(row=row, column=col_uni).value or "pcs").strip()
            
            col_cj = cols_cfg.get("cant_caja")
            cant_caja_val = 1
            if col_cj:
                try:
                    raw_cj = ws_cur.cell(row=row, column=col_cj).value
                    if raw_cj is not None:
                        cant_caja_val = float(str(raw_cj).replace(",", "."))
                        if cant_caja_val <= 0:
                            cant_caja_val = 1
                except Exception:
                    cant_caja_val = 1
            
            brand_theme = get_brand_theme(tipo_str)
            brand_name = brand_theme["display_name"]
            
            productos.append({
                "cod": cod_str,
                "nombre": nombre_str,
                "categoria": cat_str,
                "marca": brand_name,
                "size": size_str,
                "uni": uni_str,
                "cant_caja": cant_caja_val
            })
            
            marcas[brand_name] = marcas.get(brand_name, 0) + 1
            if brand_name not in categorias:
                categorias[brand_name] = {}
            categorias[brand_name][cat_str] = categorias[brand_name].get(cat_str, 0) + 1
            
    return {
        "productos": productos,
        "marcas": marcas,
        "categorias": categorias,
        "total": len(productos)
    }

if __name__ == "__main__":
    generar()